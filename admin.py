"""
Blueprint de administración profesional.
Maneja todas las funciones exclusivas de administradores.
"""
from models.usuario import Usuario
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file, abort
from functools import wraps
from datetime import datetime, timedelta
from pathlib import Path
import json
import os
import sys
import flask
from collections import defaultdict
from werkzeug.utils import secure_filename

# Importar modelos
from models import Solicitud, Denuncia, Usuario
from models.reportes import Reportes
from extensions import db

# Importar roles desde rol.py
from rol import tiene_permiso, Permiso, obtener_roles, permiso_requerido, solo_super_admin, admin_o_super, moderador_o_superior

# Módulo de configuración persistente
import config_manager as cfg

# Importar modelos de noticias, comentarios y logs
from models.noticia import Noticia, CategoriaNoticia
from models.like_noticia import LikeNoticia
from models.comentario_noticia import ComentarioNoticia
from models.log_actividad import LogActividad, registrar_log

# Importar modelos de contenido, transparencia y menú
from models.contenido import Contenido
from models.transparencia import Transparencia
from models.menu_item import MenuItem

# Importar para reportes avanzados
import pandas as pd
from models.reporte_guardado import ReporteGuardado

# ReportePlantilla es opcional, lo importamos con try/except
try:
    from models.plantilla_reporte import ReportePlantilla
except ImportError:
    ReportePlantilla = None

# IMPORT CORREGIDO: generar_pdf_desde_template en lugar de generar_pdf_desde_html
from utils.reportes_utils import (
    dataframe_desde_solicitudes, dataframe_desde_denuncias,
    dataframe_desde_usuarios, dataframe_desde_citas,
    dataframe_desde_contactos, generar_grafico_barras,
    generar_tabla_html_profesional, exportar_excel_profesional,
    generar_pdf_desde_template
)

# Intentar importar nombres desde app.py
try:
    from app import NOMBRES_SERVICIOS, NOMBRES_DENUNCIAS, SERVICIOS_CITAS, cache, REDIS_AVAILABLE
except ImportError:
    NOMBRES_SERVICIOS = {
        "funeraria": "Funerarias Municipales",
        "uso-suelo": "Certificado de Uso de Suelo",
        "oaim": "Oficina de Acceso a la Información (OAI/M)",
        "planeamiento": "Planeamiento Urbano",
        "ornato": "Ornato y Préstamos de Áreas",
        "catastro": "Catastro Municipal",
        "aseo-comercial": "Gestión Comercial de Aseo",
    }
    
    NOMBRES_DENUNCIAS = {
        "policia": "Policía Municipal",
        "limpieza": "Limpieza y Cuidado de la Vía Pública",
        "basura": "Recogida de Basura",
        "alumbrado": "Alumbrado Público",
        "otro": "Otra denuncia",
    }
    
    SERVICIOS_CITAS = {
        "asesoria-legal": "Asesoría Legal Municipal",
        "licencias": "Licencias de Funcionamiento",
        "catastro": "Trámites de Catastro",
        "registro-civil": "Registro Civil",
        "atencion-vecinal": "Atención Vecinal",
        "otro": "Otro trámite"
    }
    
    cache = None
    REDIS_AVAILABLE = False

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


# ================================================================
# FUNCIONES DE NOTIFICACIONES (sin cambios)
# ================================================================

from models.notificacion import Notificacion

def enviar_notificacion_solicitud(solicitud, estado_anterior, comentario=None):
    titulo = f"Actualización de tu solicitud {solicitud.folio}"
    
    if estado_anterior != solicitud.estado:
        mensaje = f"Tu solicitud ha cambiado de estado: **{estado_anterior}** → **{solicitud.estado}**"
    elif comentario:
        mensaje = f"El administrador ha respondido a tu solicitud: {comentario[:200]}"
    else:
        mensaje = f"Tu solicitud ha sido actualizada. Nuevo estado: {solicitud.estado}"
    
    datos_extra = {
        'folio': solicitud.folio,
        'tipo': 'solicitud',
        'estado': solicitud.estado,
        'url': url_for('mis_tramites')
    }
    
    Notificacion.crear_notificacion(
        usuario_email=solicitud.usuario_email,
        tipo='solicitud',
        titulo=titulo,
        mensaje=mensaje,
        datos_extra=datos_extra
    )


def enviar_notificacion_denuncia(denuncia, estado_anterior, comentario=None):
    titulo = f"Actualización de tu denuncia {denuncia.folio}"
    
    if estado_anterior != denuncia.estado:
        mensaje = f"Tu denuncia ha cambiado de estado: **{estado_anterior}** → **{denuncia.estado}**"
    elif comentario:
        mensaje = f"El administrador ha respondido a tu denuncia: {comentario[:200]}"
    else:
        mensaje = f"Tu denuncia ha sido actualizada. Nuevo estado: {denuncia.estado}"
    
    datos_extra = {
        'folio': denuncia.folio,
        'tipo': 'denuncia',
        'estado': denuncia.estado,
        'url': url_for('mis_tramites')
    }
    
    Notificacion.crear_notificacion(
        usuario_email=denuncia.usuario_email if not getattr(denuncia, 'anonimo', False) else None,
        tipo='denuncia',
        titulo=titulo,
        mensaje=mensaje,
        datos_extra=datos_extra
    )


def enviar_notificacion_cita(cita, estado_anterior, notas=None):
    titulo = f"Actualización de tu cita {cita.folio}"
    
    if estado_anterior != cita.estado:
        mensaje = f"Tu cita ha cambiado de estado: **{estado_anterior}** → **{cita.estado}**"
    elif notas:
        mensaje = f"El administrador ha agregado notas a tu cita: {notas[:200]}"
    else:
        mensaje = f"Tu cita ha sido actualizada. Nuevo estado: {cita.estado}"
    
    datos_extra = {
        'folio': cita.folio,
        'tipo': 'cita',
        'estado': cita.estado,
        'url': url_for('mis_citas')
    }
    
    Notificacion.crear_notificacion(
        usuario_email=cita.usuario_email,
        tipo='cita',
        titulo=titulo,
        mensaje=mensaje,
        datos_extra=datos_extra
    )


# ================================================================
# FUNCIONES AUXILIARES (sin cambios)
# ================================================================

def _cargar_usuarios_dict():
    usuarios = Usuario.query.all()
    usuarios_dict = {}
    for u in usuarios:
        usuarios_dict[u.email] = {
            "password": u.password,
            "nombre": u.nombre or "",
            "apellidos": u.apellidos or "",
            "nombre_completo": u.nombre_completo or "",
            "email": u.email,
            "tipo": u.tipo,
            "rol": u.rol,
            "telefono": u.telefono or "",
            "cedula": u.cedula or "",
            "fecha_nacimiento": u.fecha_nacimiento or "",
            "direccion": u.direccion or "",
            "foto_perfil": u.foto_perfil or "",
            "activo": u.activo,
            "fecha_registro": u.fecha_registro.strftime("%d/%m/%Y") if u.fecha_registro else "",
            "ultimo_acceso": u.ultimo_acceso.isoformat() if u.ultimo_acceso else None,
            "notas_admin": u.notas_admin or ""
        }
    return usuarios_dict


def _guardar_usuarios_db(usuarios_dict):
    for email, datos in usuarios_dict.items():
        usuario = Usuario.query.filter_by(email=email).first()
        if usuario:
            usuario.nombre = datos.get("nombre", "")
            usuario.apellidos = datos.get("apellidos", "")
            usuario.nombre_completo = datos.get("nombre_completo", "")
            usuario.tipo = datos.get("tipo", "ciudadano")
            usuario.rol = datos.get("rol")
            usuario.telefono = datos.get("telefono", "")
            usuario.cedula = datos.get("cedula", "")
            usuario.direccion = datos.get("direccion", "")
            usuario.foto_perfil = datos.get("foto_perfil", "")
            usuario.activo = datos.get("activo", True)
            usuario.notas_admin = datos.get("notas_admin", "")
            if datos.get("password"):
                usuario.password = datos["password"]
        else:
            nuevo = Usuario(
                email=email,
                password=datos.get("password", ""),
                nombre=datos.get("nombre", ""),
                apellidos=datos.get("apellidos", ""),
                nombre_completo=datos.get("nombre_completo", ""),
                tipo=datos.get("tipo", "ciudadano"),
                rol=datos.get("rol"),
                telefono=datos.get("telefono", ""),
                cedula=datos.get("cedula", ""),
                direccion=datos.get("direccion", ""),
                foto_perfil=datos.get("foto_perfil", ""),
                activo=datos.get("activo", True),
                notas_admin=datos.get("notas_admin", "")
            )
            db.session.add(nuevo)
    db.session.commit()


def formatear_fecha_para_template(fecha):
    if not fecha:
        return ""
    if isinstance(fecha, datetime):
        return fecha.strftime('%Y-%m-%d')
    if isinstance(fecha, str):
        return fecha[:10] if len(fecha) >= 10 else fecha
    return ""


def agregar_fecha_formateada(objeto):
    if hasattr(objeto, 'fecha_creacion'):
        objeto.fecha_str = formatear_fecha_para_template(objeto.fecha_creacion)
    return objeto


def sanitizar_comentarios_admin(denuncia):
    comentarios = denuncia.comentarios_admin or []
    for c in comentarios:
        if isinstance(c.get('fecha'), datetime):
            c['fecha'] = c['fecha'].isoformat()
    denuncia.comentarios_admin = comentarios
    return denuncia


# ================================================================
# DECORADOR ESPECÍFICO PARA ADMIN
# ================================================================

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            flash("Acceso restringido a administradores.", "error")
            return redirect(url_for("auth.login"))
        
        user_email = session.get("user")
        usuario = Usuario.query.filter_by(email=user_email).first()
        
        if not usuario:
            flash("Usuario no encontrado.", "error")
            return redirect(url_for("auth.login"))
        
        user_rol = usuario.rol
        
        if user_rol not in ["super_admin", "admin", "moderador"] and usuario.tipo != "admin":
            flash("Acceso restringido a administradores.", "error")
            return redirect(url_for("index"))
            
        session["user_rol"] = user_rol
        return f(*args, **kwargs)
    return decorated


# ================================================================
# DASHBOARD PRINCIPAL (CORREGIDO: SE AGREGA ENCUESTAS_STATS)
# ================================================================

@admin_bp.route("/")
@admin_bp.route("/dashboard")
@admin_required
@moderador_o_superior
def dashboard():
    try:
        solicitudes = Solicitud.query.all()
        denuncias = Denuncia.query.all()
        usuarios = Usuario.query.all()
        
        total_usuarios = len(usuarios)
        admins = 0
        for u in usuarios:
            if u.tipo == 'admin' or u.rol in ['super_admin', 'admin', 'moderador']:
                admins += 1
        ciudadanos = total_usuarios - admins
        
        total_solicitudes = len(solicitudes)
        solicitudes_pendientes = len([s for s in solicitudes if s.estado in ['pendiente', 'en_proceso']])
        solicitudes_completadas = len([s for s in solicitudes if s.estado == 'completado'])
        solicitudes_por_servicio = defaultdict(int)
        for s in solicitudes:
            servicio_id = str(s.servicio_id)
            solicitudes_por_servicio[servicio_id] += 1
        
        total_denuncias = len(denuncias)
        denuncias_pendientes = len([d for d in denuncias if d.estado in ['pendiente', 'en_investigacion']])
        denuncias_resueltas = len([d for d in denuncias if d.estado == 'resuelto'])
        denuncias_por_tipo = defaultdict(int)
        for d in denuncias:
            tipo = d.tipo
            denuncias_por_tipo[tipo] += 1
        
        hoy = datetime.now().strftime("%Y-%m-%d")
        
        solicitudes_hoy = 0
        for s in solicitudes:
            if s.fecha_creacion:
                try:
                    if isinstance(s.fecha_creacion, str):
                        fecha_str = s.fecha_creacion[:10] if len(s.fecha_creacion) >= 10 else ""
                    else:
                        fecha_str = s.fecha_creacion.strftime('%Y-%m-%d')
                    if fecha_str == hoy:
                        solicitudes_hoy += 1
                except:
                    pass
        
        denuncias_hoy = 0
        for d in denuncias:
            if d.fecha_creacion:
                try:
                    if isinstance(d.fecha_creacion, str):
                        fecha_str = d.fecha_creacion[:10] if len(d.fecha_creacion) >= 10 else ""
                    else:
                        fecha_str = d.fecha_creacion.strftime('%Y-%m-%d')
                    if fecha_str == hoy:
                        denuncias_hoy += 1
                except:
                    pass
        
        from datetime import timedelta
        solicitudes_por_mes = {}
        for i in range(5, -1, -1):
            fecha = datetime.now() - timedelta(days=30*i)
            mes = fecha.strftime("%Y-%m")
            count = 0
            for s in solicitudes:
                if s.fecha_creacion:
                    try:
                        if isinstance(s.fecha_creacion, str):
                            fecha_mes = s.fecha_creacion[:7] if len(s.fecha_creacion) >= 7 else ""
                        else:
                            fecha_mes = s.fecha_creacion.strftime('%Y-%m')
                        if fecha_mes == mes:
                            count += 1
                    except:
                        pass
            solicitudes_por_mes[mes] = count
        
        stats = {
            'usuarios': {
                'total': total_usuarios,
                'admins': admins,
                'ciudadanos': ciudadanos
            },
            'solicitudes': {
                'total': total_solicitudes,
                'pendientes': solicitudes_pendientes,
                'completadas': solicitudes_completadas,
                'por_servicio': dict(solicitudes_por_servicio),
                'por_mes': solicitudes_por_mes
            },
            'denuncias': {
                'total': total_denuncias,
                'pendientes': denuncias_pendientes,
                'resueltas': denuncias_resueltas,
                'por_tipo': dict(denuncias_por_tipo)
            },
            'actividad_hoy': {
                'total': solicitudes_hoy + denuncias_hoy,
                'solicitudes': solicitudes_hoy,
                'denuncias': denuncias_hoy
            }
        }

        def get_fecha(obj):
            if obj.fecha_creacion:
                if isinstance(obj.fecha_creacion, str):
                    return obj.fecha_creacion
                return obj.fecha_creacion.isoformat()
            return ""
        
        ultimas_solicitudes = sorted(solicitudes, key=get_fecha, reverse=True)[:5]
        ultimas_denuncias = sorted(denuncias, key=get_fecha, reverse=True)[:5]

        # ========== CORRECCIÓN: CITAS Y ENCUESTAS ==========
        try:
            from models.cita import Cita
            from models.encuesta import Encuesta   # ← Importante
            citas = Cita.query.all()
            citas_pendientes = len([c for c in citas if c.estado == 'pendiente'])
            encuestas_stats = Encuesta.obtener_estadisticas()
        except Exception as e:
            print(f"Error cargando citas/encuestas: {e}")
            citas_pendientes = 0
            encuestas_stats = {
                'total': 0, 'promedio': 0,
                'por_calificacion': {1:0, 2:0, 3:0, 4:0, 5:0},
                'por_tipo': {'solicitud': 0, 'denuncia': 0, 'cita': 0},
                'ultimas': []
            }

    except Exception as e:
        print(f"Error en dashboard: {e}")
        import traceback
        traceback.print_exc()
        stats = {
            'usuarios': {'total': 0, 'admins': 0, 'ciudadanos': 0},
            'solicitudes': {'total': 0, 'pendientes': 0, 'completadas': 0, 'por_servicio': {}, 'por_mes': {}},
            'denuncias': {'total': 0, 'pendientes': 0, 'resueltas': 0, 'por_tipo': {}},
            'actividad_hoy': {'total': 0, 'solicitudes': 0, 'denuncias': 0}
        }
        ultimas_solicitudes = []
        ultimas_denuncias = []
        citas_pendientes = 0
        encuestas_stats = {
            'total': 0, 'promedio': 0,
            'por_calificacion': {1:0, 2:0, 3:0, 4:0, 5:0},
            'por_tipo': {'solicitud': 0, 'denuncia': 0, 'cita': 0},
            'ultimas': []
        }

    return render_template(
        "admin/dashboard.html",
        stats=stats,
        ultimas_solicitudes=ultimas_solicitudes,
        ultimas_denuncias=ultimas_denuncias,
        servicios=NOMBRES_SERVICIOS,
        tipos_denuncia=NOMBRES_DENUNCIAS,
        citas_pendientes=citas_pendientes,
        encuestas_stats=encuestas_stats,   # ← NUEVA VARIABLE
        now=datetime.now()
    )


# ================================================================
# GESTIÓN DE SOLICITUDES (sin cambios)
# ================================================================

@admin_bp.route("/solicitudes")
@admin_required
@permiso_requerido(Permiso.VER_SOLICITUDES)
def listar_solicitudes():
    try:
        solicitudes = Solicitud.query.all()
    except:
        solicitudes = []
    
    total = len(solicitudes)
    pendientes = len([s for s in solicitudes if s.estado in ['pendiente', 'en_proceso']])
    completadas = len([s for s in solicitudes if s.estado == 'completado'])
    
    stats = {
        'solicitudes': total,
        'pendientes': pendientes,
        'completados': completadas,
        'en_proceso': len([s for s in solicitudes if s.estado == 'en_proceso'])
    }
    
    def get_fecha(obj):
        if obj.fecha_creacion:
            if isinstance(obj.fecha_creacion, str):
                return obj.fecha_creacion
            return obj.fecha_creacion.isoformat()
        return ""
    
    solicitudes.sort(key=get_fecha, reverse=True)
    
    import json as _json
    solicitudes_json = _json.dumps([s.to_dict() for s in solicitudes], default=str, ensure_ascii=False)
    
    return render_template(
        "admin/solicitudes.html",
        solicitudes=solicitudes,
        stats=stats,
        estados=Solicitud.ESTADOS,
        servicios=NOMBRES_SERVICIOS,
        solicitudes_json=solicitudes_json
    )


@admin_bp.route("/solicitudes/<int:solicitud_id>")
@admin_required
@permiso_requerido(Permiso.VER_SOLICITUDES)
def detalle_solicitud(solicitud_id):
    try:
        solicitud = Solicitud.query.get(solicitud_id)
        if not solicitud:
            flash("Solicitud no encontrada.", "error")
            return redirect(url_for("admin.listar_solicitudes"))
        
        categorias = {
            'solicitud': 'Solicitudes',
            'denuncia': 'Denuncias',
            'cita': 'Citas',
            'general': 'General'
        }
        
        try:
            from models.plantilla import Plantilla
            plantillas_disponibles = Plantilla.query.filter_by(categoria='solicitud', activa=True).all()
        except:
            plantillas_disponibles = []
        
        return render_template(
            "admin/solicitud_detalle.html",
            solicitud=solicitud,
            servicios=NOMBRES_SERVICIOS,
            estados=Solicitud.ESTADOS,
            categorias=categorias,
            plantillas=plantillas_disponibles,
            now=datetime.now()
        )
    except Exception as e:
        flash(f"Error al cargar solicitud: {str(e)}", "error")
        return redirect(url_for("admin.listar_solicitudes"))


# ================================================================
# CORRECCIÓN: ACTUALIZAR SOLICITUD (MANEJO DE COMENTARIO SIN ESTADO)
# ================================================================

@admin_bp.route("/solicitudes/<int:solicitud_id>/actualizar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_SOLICITUDES)
def actualizar_solicitud(solicitud_id):
    try:
        solicitud = Solicitud.query.get(solicitud_id)
        if not solicitud:
            flash("Solicitud no encontrada.", "error")
            return redirect(url_for("admin.listar_solicitudes"))
        
        estado_anterior = solicitud.estado
        nuevo_estado = request.form.get('estado')
        comentario = request.form.get('comentario', '')
        admin_email = session.get('user')
        accion = request.form.get('accion', '')

        # Si es solo comentario sin cambio de estado
        if accion == 'comentar' or not nuevo_estado:
            if comentario:
                from sqlalchemy.orm.attributes import flag_modified
                comentarios_actuales = list(solicitud.comentarios_admin or [])
                comentarios_actuales.append({
                    'fecha': datetime.now().isoformat(),
                    'admin': admin_email,
                    'comentario': comentario
                })
                solicitud.comentarios_admin = comentarios_actuales
                flag_modified(solicitud, 'comentarios_admin')
                db.session.commit()
                flash("Comentario agregado correctamente.", "success")
                enviar_notificacion_solicitud(solicitud, estado_anterior, comentario)
            else:
                flash("El comentario no puede estar vacío.", "error")
        elif nuevo_estado in Solicitud.ESTADOS:
            solicitud.actualizar_estado(nuevo_estado, comentario, admin_email)
            flash(f"Solicitud actualizada a: {nuevo_estado}", "success")
            registrar_accion('actualizar_solicitud', f"Solicitud {solicitud.folio} actualizada a {nuevo_estado}")
            enviar_notificacion_solicitud(solicitud, estado_anterior, comentario)
        else:
            flash("Estado no válido.", "error")
        
        return redirect(url_for("admin.detalle_solicitud", solicitud_id=solicitud.id))
    except Exception as e:
        flash(f"Error al actualizar: {str(e)}", "error")
        return redirect(url_for("admin.listar_solicitudes"))


@admin_bp.route("/solicitudes/<int:solicitud_id>/eliminar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.ELIMINAR_SOLICITUDES)
def eliminar_solicitud(solicitud_id):
    try:
        solicitud = Solicitud.query.get(solicitud_id)
        if solicitud:
            folio = solicitud.folio
            db.session.delete(solicitud)
            db.session.commit()
            flash(f"Solicitud {folio} eliminada correctamente.", "success")
            registrar_accion('eliminar_solicitud', f"Solicitud {folio} eliminada")
        else:
            flash("Solicitud no encontrada.", "error")
    except Exception as e:
        flash(f"Error al eliminar: {str(e)}", "error")
    
    return redirect(url_for("admin.listar_solicitudes"))


# ================================================================
# GESTIÓN DE DENUNCIAS (sin cambios)
# ================================================================

@admin_bp.route("/denuncias")
@admin_required
@permiso_requerido(Permiso.VER_DENUNCIAS)
def listar_denuncias():
    try:
        denuncias = Denuncia.query.all()
        for d in denuncias:
            sanitizar_comentarios_admin(d)
    except:
        denuncias = []
    
    total = len(denuncias)
    pendientes = len([d for d in denuncias if d.estado in ['pendiente', 'en_investigacion']])
    resueltas = len([d for d in denuncias if d.estado == 'resuelto'])
    
    stats = {
        'denuncias': total,
        'pendientes': pendientes,
        'resueltas': resueltas,
        'en_investigacion': len([d for d in denuncias if d.estado == 'en_investigacion'])
    }
    
    def get_fecha(obj):
        if obj.fecha_creacion:
            if isinstance(obj.fecha_creacion, str):
                return obj.fecha_creacion
            return obj.fecha_creacion.isoformat()
        return ""
    
    denuncias.sort(key=get_fecha, reverse=True)
    
    filtros = {
        'estado': request.args.get('estado', ''),
        'tipo': request.args.get('tipo', ''),
        'fecha_inicio': request.args.get('fecha_inicio', ''),
        'fecha_fin': request.args.get('fecha_fin', ''),
        'busqueda': request.args.get('busqueda', '')
    }
    
    import json as _json
    denuncias_json = _json.dumps([d.to_dict() for d in denuncias], default=str, ensure_ascii=False)
    
    return render_template(
        "admin/denuncias.html",
        denuncias=denuncias,
        stats=stats,
        estados=Denuncia.ESTADOS,
        tipos=NOMBRES_DENUNCIAS,
        filtros=filtros,
        denuncias_json=denuncias_json
    )


@admin_bp.route("/denuncias/<int:denuncia_id>")
@admin_required
@permiso_requerido(Permiso.VER_DENUNCIAS)
def detalle_denuncia(denuncia_id):
    try:
        denuncia = Denuncia.query.get(denuncia_id)
        if not denuncia:
            flash("Denuncia no encontrada.", "error")
            return redirect(url_for("admin.listar_denuncias"))
        
        denuncia = sanitizar_comentarios_admin(denuncia)
        
        categorias = {
            'solicitud': 'Solicitudes',
            'denuncia': 'Denuncias',
            'cita': 'Citas',
            'general': 'General'
        }
        
        try:
            from models.plantilla import Plantilla
            plantillas_disponibles = Plantilla.query.filter_by(categoria='denuncia', activa=True).all()
        except:
            plantillas_disponibles = []
        
        return render_template(
            "admin/denuncia_detalle.html",
            denuncia=denuncia,
            tipos=NOMBRES_DENUNCIAS,
            estados=Denuncia.ESTADOS,
            categorias=categorias,
            plantillas=plantillas_disponibles,
            now=datetime.now()
        )
    except Exception as e:
        flash(f"Error al cargar denuncia: {str(e)}", "error")
        return redirect(url_for("admin.listar_denuncias"))


@admin_bp.route("/denuncias/<int:denuncia_id>/actualizar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_DENUNCIAS)
def actualizar_denuncia(denuncia_id):
    try:
        denuncia = Denuncia.query.get(denuncia_id)
        if not denuncia:
            flash("Denuncia no encontrada.", "error")
            return redirect(url_for("admin.listar_denuncias"))
        
        estado_anterior = denuncia.estado
        nuevo_estado = request.form.get('estado')
        comentario = request.form.get('comentario', '')
        admin_email = session.get('user')
        
        if nuevo_estado in Denuncia.ESTADOS:
            denuncia.estado = nuevo_estado
            denuncia.fecha_actualizacion = datetime.utcnow()
            
            if comentario:
                from sqlalchemy.orm.attributes import flag_modified
                comentarios_actuales = list(denuncia.comentarios_admin or [])
                comentarios_actuales.append({
                    'fecha': datetime.now().isoformat(),
                    'admin': admin_email,
                    'comentario': comentario
                })
                denuncia.comentarios_admin = comentarios_actuales
                flag_modified(denuncia, 'comentarios_admin')
            
            db.session.commit()
            
            flash(f"Denuncia actualizada a: {nuevo_estado}", "success")
            registrar_accion('actualizar_denuncia', f"Denuncia {denuncia.folio} actualizada a {nuevo_estado}")
            
            if not getattr(denuncia, 'anonimo', False):
                enviar_notificacion_denuncia(denuncia, estado_anterior, comentario)
        else:
            flash("Estado no válido.", "error")
        
        return redirect(url_for("admin.detalle_denuncia", denuncia_id=denuncia.id))
    except Exception as e:
        db.session.rollback()
        flash(f"Error al actualizar: {str(e)}", "error")
        return redirect(url_for("admin.listar_denuncias"))


@admin_bp.route("/denuncias/<int:denuncia_id>/eliminar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.ELIMINAR_DENUNCIAS)
def eliminar_denuncia(denuncia_id):
    try:
        denuncia = Denuncia.query.get(denuncia_id)
        if denuncia:
            folio = denuncia.folio
            db.session.delete(denuncia)
            db.session.commit()
            flash(f"Denuncia {folio} eliminada correctamente.", "success")
            registrar_accion('eliminar_denuncia', f"Denuncia {folio} eliminada")
        else:
            flash("Denuncia no encontrada.", "error")
    except Exception as e:
        flash(f"Error al eliminar: {str(e)}", "error")
    
    return redirect(url_for("admin.listar_denuncias"))


# ================================================================
# GESTIÓN DE USUARIOS (CORREGIDO: se agrega stats al render_template)
# ================================================================

@admin_bp.route("/usuarios")
@admin_required
@permiso_requerido(Permiso.VER_USUARIOS)
def listar_usuarios():
    busqueda = request.args.get('q', '')
    tipo = request.args.get('tipo', '')
    rol = request.args.get('rol', '')

    query = Usuario.query
    
    if tipo:
        query = query.filter_by(tipo=tipo)
    if rol:
        query = query.filter_by(rol=rol)
    if busqueda:
        b = busqueda.lower()
        query = query.filter(
            db.or_(
                Usuario.email.ilike(f'%{b}%'),
                Usuario.nombre.ilike(f'%{b}%')
            )
        )
    
    usuarios = query.all()
    
    usuarios_lista = []
    for u in usuarios:
        usuarios_lista.append({
            "email": u.email,
            "nombre": u.nombre or "",
            "apellidos": u.apellidos or "",
            "telefono": u.telefono or "",
            "tipo": u.tipo,
            "rol": u.rol,
            "activo": u.activo,
            "fecha_registro": u.fecha_registro.strftime("%d/%m/%Y") if u.fecha_registro else "",
            "ultimo_acceso": u.ultimo_acceso.isoformat() if u.ultimo_acceso else None,
            "notas_admin": u.notas_admin or ""
        })

    # ========== CORRECCIÓN: AÑADIR stats ==========
    # Estadísticas globales (sin aplicar filtros de búsqueda/tipo/rol)
    stats = {
        'total': Usuario.query.count(),
        'activos': Usuario.query.filter_by(activo=True).count(),
        'inactivos': Usuario.query.filter_by(activo=False).count(),
        'admins': Usuario.query.filter(Usuario.rol.in_(['super_admin', 'admin', 'moderador'])).count(),
        'ciudadanos': Usuario.query.filter(Usuario.tipo == 'ciudadano').count()
    }

    return render_template(
        "admin/usuarios.html",
        usuarios=usuarios_lista,
        roles=obtener_roles(),
        filtros={'q': busqueda, 'tipo': tipo, 'rol': rol},
        stats=stats   # ← SE PASA stats AL TEMPLATE
    )


@admin_bp.route("/usuarios/<path:email>")
@admin_required
@permiso_requerido(Permiso.VER_USUARIOS)
def detalle_usuario(email):
    usuario = Usuario.query.filter_by(email=email).first()
    
    if not usuario:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("admin.listar_usuarios"))

    usuario_dict = {
        "email": usuario.email,
        "nombre": usuario.nombre or "",
        "apellidos": usuario.apellidos or "",
        "telefono": usuario.telefono or "",
        "tipo": usuario.tipo,
        "rol": usuario.rol,
        "activo": usuario.activo,
        "fecha_registro": usuario.fecha_registro.strftime("%d/%m/%Y") if usuario.fecha_registro else "",
        "ultimo_acceso": usuario.ultimo_acceso.isoformat() if usuario.ultimo_acceso else None,
        "notas_admin": usuario.notas_admin or ""
    }

    try:
        solicitudes = Solicitud.query.filter_by(usuario_email=email).all()
        denuncias = Denuncia.query.filter_by(usuario_email=email).all()
    except:
        solicitudes = []
        denuncias = []

    return render_template(
        "admin/usuario_detalle.html",
        usuario=usuario_dict,
        roles=obtener_roles(),
        solicitudes=solicitudes,
        denuncias=denuncias
    )


@admin_bp.route("/usuarios/<path:email>/editar", methods=["GET", "POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_USUARIOS)
def editar_usuario(email):
    usuario = Usuario.query.filter_by(email=email).first()
    
    if not usuario:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("admin.listar_usuarios"))

    if request.method == "POST":
        usuario.nombre = request.form.get('nombre', '')
        usuario.apellidos = request.form.get('apellidos', '')
        usuario.telefono = request.form.get('telefono', '')
        usuario.tipo = request.form.get('tipo', 'ciudadano')
        usuario.activo = request.form.get('activo') == 'on'
        usuario.notas_admin = request.form.get('notas_admin', '')
        
        if session.get("user_rol") == "super_admin":
            nuevo_rol = request.form.get('rol')
            if nuevo_rol in ["super_admin", "admin", "moderador", ""]:
                usuario.rol = nuevo_rol if nuevo_rol else None

        db.session.commit()
        
        flash(f"Usuario {email} actualizado correctamente.", "success")
        registrar_accion('editar_usuario', f"Usuario {email} actualizado")
        return redirect(url_for("admin.detalle_usuario", email=email))

    usuario_dict = {
        "email": usuario.email,
        "nombre": usuario.nombre or "",
        "apellidos": usuario.apellidos or "",
        "telefono": usuario.telefono or "",
        "tipo": usuario.tipo,
        "rol": usuario.rol,
        "activo": usuario.activo,
        "fecha_registro": usuario.fecha_registro.strftime("%d/%m/%Y") if usuario.fecha_registro else "",
        "notas_admin": usuario.notas_admin or ""
    }

    return render_template(
        "admin/usuario_editar.html", 
        usuario=usuario_dict,
        roles=obtener_roles()
    )


@admin_bp.route("/usuarios/<path:email>/cambiar-password", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_USUARIOS)
def cambiar_password_usuario(email):
    usuario = Usuario.query.filter_by(email=email).first()
    
    if not usuario:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("admin.listar_usuarios"))

    nueva_password = request.form.get('nueva_password')
    confirmar = request.form.get('confirmar_password')

    if not nueva_password or len(nueva_password) < 6:
        flash("La contraseña debe tener al menos 6 caracteres.", "error")
        return redirect(url_for("admin.detalle_usuario", email=email))

    if nueva_password != confirmar:
        flash("Las contraseñas no coinciden.", "error")
        return redirect(url_for("admin.detalle_usuario", email=email))

    usuario.password = nueva_password
    db.session.commit()
    
    flash(f"Contraseña de {email} actualizada correctamente.", "success")
    registrar_accion('cambiar_password', f"Contraseña cambiada para {email}")
    return redirect(url_for("admin.detalle_usuario", email=email))


@admin_bp.route("/usuarios/crear-admin", methods=["GET", "POST"])
@admin_required
@solo_super_admin
def crear_admin():
    if request.method == "POST":
        nombre = request.form.get('nombre', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirmar = request.form.get('confirmar_password', '')
        telefono = request.form.get('telefono', '')
        rol = request.form.get('rol', 'moderador')

        errores = []
        if not nombre or not email or not password:
            errores.append("Todos los campos son obligatorios.")
        if password != confirmar:
            errores.append("Las contraseñas no coinciden.")
        if len(password) < 6:
            errores.append("La contraseña debe tener al menos 6 caracteres.")
        
        existe = Usuario.query.filter_by(email=email).first()
        if existe:
            errores.append("Este correo ya está registrado.")

        if errores:
            for e in errores:
                flash(e, "error")
            return render_template("admin/crear_admin.html", roles=obtener_roles())

        nuevo_usuario = Usuario(
            email=email,
            password=password,
            nombre=nombre,
            apellidos="",
            nombre_completo=nombre,
            tipo="admin",
            rol=rol,
            telefono=telefono,
            activo=True,
            notas_admin=f"Creado por {session.get('user_name', session.get('user', 'desconocido'))}"
        )

        db.session.add(nuevo_usuario)
        db.session.commit()

        rol_nombre = {
            "super_admin": "Super Administrador",
            "admin": "Administrador",
            "moderador": "Moderador"
        }.get(rol, rol)

        flash(f"{rol_nombre} {email} creado exitosamente.", "success")
        registrar_accion('crear_admin', f"{rol_nombre} {email} creado")
        return redirect(url_for("admin.listar_usuarios"))

    return render_template("admin/crear_admin.html", roles=obtener_roles())


@admin_bp.route("/usuarios/<path:email>/toggle-activo", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_USUARIOS)
def toggle_usuario_activo(email):
    usuario = Usuario.query.filter_by(email=email).first()
    
    if not usuario:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("admin.listar_usuarios"))

    if email == session.get('user'):
        flash("No puedes desactivar tu propia cuenta.", "error")
        return redirect(url_for("admin.detalle_usuario", email=email))

    estado_actual = usuario.activo
    usuario.activo = not estado_actual
    db.session.commit()
    
    nuevo_estado = "activado" if not estado_actual else "desactivado"
    flash(f"Usuario {email} {nuevo_estado}.", "success")
    registrar_accion('toggle_activo', f"Usuario {email} {nuevo_estado}")
    return redirect(url_for("admin.detalle_usuario", email=email))


# ================================================================
# REPORTES Y ESTADÍSTICAS (sin cambios)
# ================================================================

@admin_bp.route("/reportes")
@admin_required
@permiso_requerido(Permiso.VER_BITACORA)
def reportes():
    try:
        stats = Reportes.obtener_estadisticas_generales()
        solicitudes_por_servicio = Reportes.obtener_solicitudes_por_servicio()
        denuncias_por_tipo = Reportes.obtener_denuncias_por_tipo()
    except:
        stats = {}
        solicitudes_por_servicio = {}
        denuncias_por_tipo = {}

    return render_template(
        "admin/reportes.html",
        stats=stats,
        solicitudes_por_servicio=solicitudes_por_servicio,
        denuncias_por_tipo=denuncias_por_tipo,
        servicios=NOMBRES_SERVICIOS,
        tipos_denuncia=NOMBRES_DENUNCIAS
    )


# ================================================================
# CITAS (sin cambios)
# ================================================================

@admin_bp.route("/citas")
@admin_required
def admin_citas():
    try:
        from models.cita import Cita
        citas = Cita.query.all()
        
        def get_fecha(c):
            if c.fecha:
                if hasattr(c.fecha, 'strftime'):
                    return c.fecha.strftime('%Y-%m-%d') + ' ' + (c.hora or '')
                else:
                    return str(c.fecha) + ' ' + (c.hora or '')
            return (c.hora or '')
        
        citas.sort(key=get_fecha)
        
        hoy = datetime.now().strftime('%Y-%m-%d')
        stats = {
            'total': len(citas),
            'pendientes': len([c for c in citas if c.estado == 'pendiente']),
            'confirmadas': len([c for c in citas if c.estado == 'confirmada']),
            'hoy': len([c for c in citas if c.fecha and str(c.fecha) == hoy])
        }
        
        return render_template("admin/citas.html", citas=citas, stats=stats, servicios=SERVICIOS_CITAS)
    except Exception as e:
        flash(f"Error al cargar citas: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


@admin_bp.route("/citas/<int:cita_id>/estado", methods=["POST"])
@admin_required
def admin_cambiar_estado_cita(cita_id):
    try:
        from models.cita import Cita
        cita = Cita.query.get(cita_id)
        if not cita:
            flash("❌ Cita no encontrada.", "error")
            return redirect(url_for('admin.admin_citas'))
        
        estado_anterior = cita.estado
        nuevo_estado = request.form.get("estado")
        notas = request.form.get("notas", "")
        
        if nuevo_estado not in Cita.ESTADOS:
            flash("❌ Estado no válido.", "error")
            return redirect(url_for('admin.admin_citas'))
        
        cita.estado = nuevo_estado
        if notas:
            cita.notas_admin = notas
        
        db.session.commit()
        
        flash(f"✅ Estado de cita actualizado a: {nuevo_estado}", "success")
        
        enviar_notificacion_cita(cita, estado_anterior, notas)
        
        return redirect(url_for('admin.admin_citas'))
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


# ================================================================
# PLANTILLAS (con corrección: estilos={})
# ================================================================

@admin_bp.route("/plantillas")
@admin_required
def admin_plantillas():
    try:
        from models.plantilla import Plantilla
        plantillas = Plantilla.query.all()
        
        por_categoria = {}
        for p in plantillas:
            if p.categoria not in por_categoria:
                por_categoria[p.categoria] = []
            por_categoria[p.categoria].append(p)
        
        # ✅ CORRECCIÓN: se agregó estilos={}
        return render_template("admin/plantillas.html", 
                               plantillas=plantillas, 
                               por_categoria=por_categoria, 
                               categorias=Plantilla.CATEGORIAS,
                               estilos={})
    except Exception as e:
        flash(f"Error al cargar plantillas: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


@admin_bp.route("/plantillas/crear", methods=["POST"])
@admin_required
def admin_crear_plantilla():
    try:
        from models.plantilla import Plantilla
        nombre = request.form.get("nombre")
        categoria = request.form.get("categoria")
        contenido = request.form.get("contenido")
        variables = request.form.getlist("variables") or ['folio', 'nombre', 'fecha']
        
        if not nombre or not contenido:
            flash("❌ Nombre y contenido son obligatorios.", "error")
            return redirect(url_for('admin.admin_plantillas'))
        
        nueva = Plantilla(
            nombre=nombre,
            categoria=categoria,
            contenido=contenido,
            variables=variables,
            creada_por=session.get("user_name", session["user"])
        )
        db.session.add(nueva)
        db.session.commit()
        
        flash(f"✅ Plantilla '{nombre}' creada correctamente.", "success")
        return redirect(url_for('admin.admin_plantillas'))
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


@admin_bp.route("/plantillas/<int:plantilla_id>/toggle", methods=["POST"])
@admin_required
def admin_toggle_plantilla(plantilla_id):
    try:
        from models.plantilla import Plantilla
        plantilla = Plantilla.query.get(plantilla_id)
        if not plantilla:
            flash("❌ Plantilla no encontrada.", "error")
            return redirect(url_for('admin.admin_plantillas'))
        
        plantilla.activa = not plantilla.activa
        db.session.commit()
        
        estado = "activada" if plantilla.activa else "desactivada"
        flash(f"✅ Plantilla '{plantilla.nombre}' {estado}.", "success")
        return redirect(url_for('admin.admin_plantillas'))
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


@admin_bp.route("/plantillas/<int:plantilla_id>/eliminar", methods=["POST"])
@admin_required
def admin_eliminar_plantilla(plantilla_id):
    try:
        from models.plantilla import Plantilla
        plantilla = Plantilla.query.get(plantilla_id)
        if plantilla:
            db.session.delete(plantilla)
            db.session.commit()
            flash("✅ Plantilla eliminada.", "success")
        else:
            flash("❌ Plantilla no encontrada.", "error")
        return redirect(url_for('admin.admin_plantillas'))
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


# ========== NUEVA RUTA API PARA LISTAR PLANTILLAS (JSON) ==========
@admin_bp.route("/api/plantillas-lista")
@admin_required
def api_plantillas_lista():
    try:
        from models.plantilla import Plantilla
        plantillas = Plantilla.query.filter_by(activa=True).all()
        return jsonify([{
            'id': p.id,
            'nombre': p.nombre,
            'contenido': p.contenido,
            'categoria': p.categoria,
            'variables': p.variables or []
        } for p in plantillas])
    except Exception as e:
        return jsonify([])


# ================================================================
# MAPA DE INCIDENCIAS (sin cambios)
# ================================================================

@admin_bp.route("/mapa")
@admin_required
def admin_mapa_incidencias():
    try:
        denuncias = Denuncia.query.all()

        denuncias_geo = []
        for d in denuncias:
            if not getattr(d, 'geolocalizada', False):
                continue
            try:
                lat = float(d.lat) if d.lat else None
                lng = float(d.lng) if d.lng else None
            except (ValueError, TypeError):
                lat, lng = None, None

            if not lat or not lng:
                continue

            denuncias_geo.append({
                'id':            d.id,
                'folio':         d.folio,
                'tipo':          d.tipo,
                'tipo_nombre':   getattr(d, 'tipo_nombre', None) or NOMBRES_DENUNCIAS.get(d.tipo, d.tipo),
                'estado':        d.estado,
                'descripcion':   (d.descripcion or '')[:200],
                'ubicacion':     d.ubicacion or '',
                'direccion':     d.ubicacion or '',
                'usuario_nombre': getattr(d, 'usuario_nombre', None) or 'Anónimo',
                'lat':           lat,
                'lng':           lng,
                'fecha':         d.fecha_creacion.isoformat()
                                 if hasattr(d.fecha_creacion, 'isoformat')
                                 else str(d.fecha_creacion or ''),
            })

        stats = {
            'total':          len(denuncias),
            'geolocalizadas': len(denuncias_geo),
            'pendientes':     len([d for d in denuncias
                                   if d.estado in ['pendiente', 'en_investigacion']]),
        }

        return render_template(
            "admin/mapa_admin.html",
            denuncias=denuncias_geo,
            tipos=NOMBRES_DENUNCIAS,
            stats=stats
        )
    except Exception as e:
        flash(f"Error al cargar mapa: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


# ================================================================
# ENCUESTAS (sin cambios)
# ================================================================

@admin_bp.route("/encuestas")
@admin_required
def admin_encuestas():
    try:
        from models.encuesta import Encuesta
        stats = Encuesta.obtener_estadisticas()
        return render_template("admin/encuestas.html", stats=stats)
    except Exception as e:
        flash(f"Error al cargar encuestas: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


# ================================================================
# GESTIÓN DE CATEGORÍAS DE NOTICIAS (INSERTADO JUSTO ANTES DE admin_noticias)
# ================================================================

@admin_bp.route("/noticias/categorias")
@admin_required
def admin_categorias():
    categorias = CategoriaNoticia.query.order_by(CategoriaNoticia.orden.asc(), CategoriaNoticia.nombre.asc()).all()
    return render_template("admin/categorias.html", categorias=categorias)


@admin_bp.route("/noticias/categorias/crear", methods=["POST"])
@admin_required
def admin_categoria_crear():
    nombre = request.form.get('nombre', '').strip()
    color  = request.form.get('color', '#3b82f6').strip()
    icono  = request.form.get('icono', 'newspaper').strip()
    orden  = request.form.get('orden', 0, type=int)

    if not nombre:
        flash("El nombre de la categoría es obligatorio.", "error")
        return redirect(url_for('admin.admin_categorias'))

    slug = CategoriaNoticia.slugify(nombre)

    if CategoriaNoticia.query.filter_by(slug=slug).first():
        flash(f"Ya existe una categoría con el nombre '{nombre}'.", "error")
        return redirect(url_for('admin.admin_categorias'))

    cat = CategoriaNoticia(nombre=nombre, slug=slug, color=color, icono=icono, orden=orden, activa=True)
    db.session.add(cat)
    db.session.commit()

    registrar_log(accion='crear_categoria', modulo='noticias', descripcion=f"Creó categoría '{nombre}'")
    flash(f"✅ Categoría '{nombre}' creada correctamente.", "success")
    return redirect(url_for('admin.admin_categorias'))


@admin_bp.route("/noticias/categorias/<int:cat_id>/editar", methods=["POST"])
@admin_required
def admin_categoria_editar(cat_id):
    cat = CategoriaNoticia.query.get_or_404(cat_id)

    nombre = request.form.get('nombre', '').strip()
    color  = request.form.get('color', cat.color).strip()
    icono  = request.form.get('icono', cat.icono).strip()
    orden  = request.form.get('orden', cat.orden, type=int)
    activa = request.form.get('activa') == 'on'

    if not nombre:
        flash("El nombre es obligatorio.", "error")
        return redirect(url_for('admin.admin_categorias'))

    nuevo_slug = CategoriaNoticia.slugify(nombre)
    existe = CategoriaNoticia.query.filter(CategoriaNoticia.slug == nuevo_slug, CategoriaNoticia.id != cat_id).first()
    if existe:
        flash(f"Ya existe otra categoría con ese nombre.", "error")
        return redirect(url_for('admin.admin_categorias'))

    cat.nombre = nombre
    cat.slug   = nuevo_slug
    cat.color  = color
    cat.icono  = icono
    cat.orden  = orden
    cat.activa = activa
    db.session.commit()

    registrar_log(accion='editar_categoria', modulo='noticias', descripcion=f"Editó categoría '{nombre}'")
    flash(f"✅ Categoría '{nombre}' actualizada.", "success")
    return redirect(url_for('admin.admin_categorias'))


@admin_bp.route("/noticias/categorias/<int:cat_id>/eliminar", methods=["POST"])
@admin_required
def admin_categoria_eliminar(cat_id):
    cat = CategoriaNoticia.query.get_or_404(cat_id)

    if cat.noticias.count() > 0:
        flash(f"No se puede eliminar '{cat.nombre}' porque tiene {cat.noticias.count()} noticia(s) asociada(s).", "error")
        return redirect(url_for('admin.admin_categorias'))

    nombre = cat.nombre
    db.session.delete(cat)
    db.session.commit()

    registrar_log(accion='eliminar_categoria', modulo='noticias', descripcion=f"Eliminó categoría '{nombre}'")
    flash(f"✅ Categoría '{nombre}' eliminada.", "success")
    return redirect(url_for('admin.admin_categorias'))


@admin_bp.route("/noticias/categorias/<int:cat_id>/toggle", methods=["POST"])
@admin_required
def admin_categoria_toggle(cat_id):
    cat = CategoriaNoticia.query.get_or_404(cat_id)
    cat.activa = not cat.activa
    db.session.commit()
    estado = "activada" if cat.activa else "desactivada"
    flash(f"✅ Categoría '{cat.nombre}' {estado}.", "success")
    return redirect(url_for('admin.admin_categorias'))


# ================================================================
# NOTICIAS - ADMINISTRACIÓN (con ruta subir-imagen ANTES de rutas con parámetros)
# ================================================================

@admin_bp.route("/noticias")
@admin_required
def admin_noticias():
    pagina = request.args.get('pagina', 1, type=int)
    estado = request.args.get('estado', '')
    categoria_id = request.args.get('categoria', type=int)
    
    query = Noticia.query
    
    if estado:
        query = query.filter_by(estado=estado)
    if categoria_id:
        query = query.filter_by(categoria_id=categoria_id)
    
    query = query.order_by(Noticia.fecha_creacion.desc())
    paginacion = query.paginate(page=pagina, per_page=20, error_out=False)
    
    categorias = CategoriaNoticia.query.order_by(CategoriaNoticia.nombre).all()
    
    return render_template(
        "admin/noticias.html",
        noticias=paginacion.items,
        paginacion=paginacion,
        categorias=categorias,
        estado_actual=estado,
        categoria_actual=categoria_id
    )


# ================================================================
# SUBIR IMAGEN DE NOTICIA A CLOUDINARY (RUTA SIN PARÁMETROS - VA PRIMERO)
# ================================================================

@admin_bp.route("/noticias/subir-imagen", methods=["POST"])
@admin_required
def admin_subir_imagen_noticia():
    """Sube una imagen de noticia a Cloudinary y devuelve la URL."""
    try:
        archivo = request.files.get('imagen')

        if not archivo or archivo.filename == '':
            return jsonify({'success': False, 'error': 'No se recibió ningún archivo'}), 400

        # Validar extensión
        ALLOWED = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        ext = archivo.filename.rsplit('.', 1)[-1].lower() if '.' in archivo.filename else ''
        if ext not in ALLOWED:
            return jsonify({'success': False, 'error': f'Formato no permitido: {ext}'}), 400

        archivo.stream.seek(0)
        file_content = archivo.read()

        if len(file_content) == 0:
            return jsonify({'success': False, 'error': 'El archivo está vacío'}), 400

        if len(file_content) > 10 * 1024 * 1024:
            return jsonify({'success': False, 'error': 'La imagen supera 10 MB'}), 400

        # Nombre único
        from datetime import datetime as _dt
        ts = _dt.now().strftime('%Y%m%d_%H%M%S')
        admin_email = session.get('user', 'admin').replace('@', '_').replace('.', '_')
        public_id = f"noticias/{admin_email}_{ts}"

        import cloudinary
        import cloudinary.uploader

        resultado = cloudinary.uploader.upload(
            file_content,
            public_id=public_id,
            folder='noticias',
            overwrite=True,
            transformation=[
                {'width': 1200, 'height': 630, 'crop': 'fill', 'gravity': 'auto'},
                {'fetch_format': 'auto', 'quality': 'auto'}
            ]
        )

        registrar_log(
            accion='subir_imagen_noticia',
            modulo='noticias',
            descripcion=f"Subió imagen de noticia: {resultado.get('public_id')}",
        )

        return jsonify({
            'success': True,
            'url': resultado['secure_url'],
            'public_id': resultado['public_id']
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ================================================================
# RUTAS CON PARÁMETROS (después de la ruta sin parámetros)
# ================================================================

@admin_bp.route("/noticias/nueva", methods=["GET", "POST"])
@admin_required
def admin_noticia_nueva():
    if request.method == "POST":
        titulo = request.form.get('titulo', '').strip()
        contenido = request.form.get('contenido', '').strip()
        resumen = request.form.get('resumen', '').strip()
        categoria_id = request.form.get('categoria_id', type=int)
        imagen_url = request.form.get('imagen_url', '').strip()
        tags = request.form.get('tags', '').strip()
        destacada = request.form.get('destacada') == 'on'
        estado = request.form.get('estado', 'borrador')
        
        errores = []
        if not titulo:
            errores.append("El título es obligatorio")
        if not contenido:
            errores.append("El contenido es obligatorio")
        if not categoria_id:
            errores.append("La categoría es obligatoria")
        
        if errores:
            for e in errores:
                flash(e, "error")
            categorias = CategoriaNoticia.query.order_by(CategoriaNoticia.nombre).all()
            return render_template("admin/noticias_form.html", categorias=categorias, noticia=None)
        
        noticia = Noticia.crear(
            titulo=titulo,
            contenido=contenido,
            autor_email=session.get('user'),
            categoria_id=categoria_id,
            resumen=resumen,
            imagen_url=imagen_url,
            tags=[t.strip() for t in tags.split(',')] if tags else [],
            destacada=destacada,
            estado=estado
        )
        
        if estado == 'publicado' and not noticia.fecha_publicacion:
            noticia.publicar()
        
        registrar_log(
            accion='crear_noticia',
            modulo='noticias',
            descripcion=f"Creó la noticia '{noticia.titulo}'",
            datos_extra={'noticia_id': noticia.id}
        )
        
        flash(f"✅ Noticia '{noticia.titulo}' creada exitosamente.", "success")
        return redirect(url_for('admin.admin_noticias'))
    
    categorias = CategoriaNoticia.query.order_by(CategoriaNoticia.nombre).all()
    return render_template("admin/noticias_form.html", categorias=categorias, noticia=None)


@admin_bp.route("/noticias/<int:noticia_id>/editar", methods=["GET", "POST"])
@admin_required
def admin_noticia_editar(noticia_id):
    noticia = Noticia.query.get_or_404(noticia_id)
    
    if request.method == "POST":
        noticia.titulo = request.form.get('titulo', '').strip()
        noticia.contenido = request.form.get('contenido', '').strip()
        noticia.resumen = request.form.get('resumen', '').strip()
        noticia.categoria_id = request.form.get('categoria_id', type=int)
        noticia.imagen_url = request.form.get('imagen_url', '').strip()
        tags = request.form.get('tags', '').strip()
        noticia.tags = [t.strip() for t in tags.split(',')] if tags else []
        noticia.destacada = request.form.get('destacada') == 'on'
        noticia.estado = request.form.get('estado', noticia.estado)
        
        db.session.commit()
        
        registrar_log(
            accion='editar_noticia',
            modulo='noticias',
            descripcion=f"Editó la noticia '{noticia.titulo}'",
            datos_extra={'noticia_id': noticia.id}
        )
        
        flash(f"✅ Noticia '{noticia.titulo}' actualizada exitosamente.", "success")
        return redirect(url_for('admin.admin_noticias'))
    
    categorias = CategoriaNoticia.query.order_by(CategoriaNoticia.nombre).all()
    return render_template("admin/noticias_form.html", categorias=categorias, noticia=noticia)


@admin_bp.route("/noticias/<int:noticia_id>/eliminar", methods=["POST"])
@admin_required
def admin_noticia_eliminar(noticia_id):
    noticia = Noticia.query.get_or_404(noticia_id)
    titulo = noticia.titulo
    
    registrar_log(
        accion='eliminar_noticia',
        modulo='noticias',
        descripcion=f"Eliminó la noticia '{titulo}'",
        datos_extra={'noticia_id': noticia_id}
    )
    
    db.session.delete(noticia)
    db.session.commit()
    
    flash(f"✅ Noticia '{titulo}' eliminada correctamente.", "success")
    return redirect(url_for('admin.admin_noticias'))


@admin_bp.route("/noticias/<int:noticia_id>/publicar", methods=["POST"])
@admin_required
def admin_noticia_publicar(noticia_id):
    noticia = Noticia.query.get_or_404(noticia_id)
    noticia.publicar()
    
    registrar_log(
        accion='publicar_noticia',
        modulo='noticias',
        descripcion=f"Publicó la noticia '{noticia.titulo}'",
        datos_extra={'noticia_id': noticia_id}
    )
    
    flash(f"✅ Noticia '{noticia.titulo}' publicada exitosamente.", "success")
    return redirect(url_for('admin.admin_noticias'))


@admin_bp.route("/noticias/<int:noticia_id>/archivar", methods=["POST"])
@admin_required
def admin_noticia_archivar(noticia_id):
    noticia = Noticia.query.get_or_404(noticia_id)
    noticia.archivar()
    
    registrar_log(
        accion='archivar_noticia',
        modulo='noticias',
        descripcion=f"Archivó la noticia '{noticia.titulo}'",
        datos_extra={'noticia_id': noticia_id}
    )
    
    flash(f"✅ Noticia '{noticia.titulo}' archivada correctamente.", "success")
    return redirect(url_for('admin.admin_noticias'))


# ================================================================
# COMENTARIOS - ADMINISTRACIÓN (sin cambios)
# ================================================================

@admin_bp.route("/noticias/comentarios")
@admin_required
def admin_comentarios():
    pagina = request.args.get('pagina', 1, type=int)
    comentarios = ComentarioNoticia.query.filter_by(aprobado=False).order_by(
        ComentarioNoticia.fecha_creacion.desc()
    ).paginate(page=pagina, per_page=20, error_out=False)
    
    return render_template("admin/comentarios.html", comentarios=comentarios)


@admin_bp.route("/noticias/comentarios/<int:comentario_id>/aprobar", methods=["POST"])
@admin_required
def admin_comentario_aprobar(comentario_id):
    comentario = ComentarioNoticia.query.get_or_404(comentario_id)
    comentario.aprobado = True
    db.session.commit()
    
    registrar_log(
        accion='aprobar_comentario',
        modulo='noticias',
        descripcion=f"Aprobó comentario de {comentario.autor_nombre}",
        datos_extra={'comentario_id': comentario_id, 'noticia_id': comentario.noticia_id}
    )
    
    flash(f"✅ Comentario de {comentario.autor_nombre} aprobado.", "success")
    return redirect(url_for('admin.admin_comentarios'))


@admin_bp.route("/noticias/comentarios/<int:comentario_id>/rechazar", methods=["POST"])
@admin_required
def admin_comentario_rechazar(comentario_id):
    comentario = ComentarioNoticia.query.get_or_404(comentario_id)
    autor_nombre = comentario.autor_nombre
    
    registrar_log(
        accion='rechazar_comentario',
        modulo='noticias',
        descripcion=f"Rechazó comentario de {autor_nombre}",
        datos_extra={'comentario_id': comentario_id, 'noticia_id': comentario.noticia_id}
    )
    
    db.session.delete(comentario)
    db.session.commit()
    
    flash(f"✅ Comentario de {autor_nombre} eliminado.", "success")
    return redirect(url_for('admin.admin_comentarios'))


# ================================================================
# LOGS - ADMINISTRACIÓN (sin cambios)
# ================================================================

@admin_bp.route("/logs")
@admin_required
@permiso_requerido(Permiso.VER_BITACORA)
def admin_logs():
    pagina = request.args.get('pagina', 1, type=int)
    
    filtros = {}
    if request.args.get('modulo'):
        filtros['modulo'] = request.args.get('modulo')
    if request.args.get('nivel'):
        filtros['nivel'] = request.args.get('nivel')
    if request.args.get('usuario'):
        filtros['usuario_email'] = request.args.get('usuario')
    if request.args.get('fecha_desde'):
        try:
            filtros['fecha_desde'] = datetime.strptime(request.args.get('fecha_desde'), '%Y-%m-%d')
        except:
            pass
    if request.args.get('fecha_hasta'):
        try:
            filtros['fecha_hasta'] = datetime.strptime(request.args.get('fecha_hasta'), '%Y-%m-%d')
        except:
            pass
    
    logs = LogActividad.listar(pagina=pagina, por_pagina=50, filtros=filtros)
    estadisticas = LogActividad.obtener_estadisticas(dias=7)
    
    modulos_disponibles = db.session.query(LogActividad.modulo).filter(
        LogActividad.modulo.isnot(None)
    ).distinct().all()
    modulos_disponibles = [m[0] for m in modulos_disponibles if m[0]]
    
    niveles_disponibles = LogActividad.NIVELES
    
    return render_template(
        "admin/logs.html",
        logs=logs,
        estadisticas=estadisticas,
        modulos_disponibles=modulos_disponibles,
        niveles_disponibles=niveles_disponibles,
        filtros_actuales=request.args
    )


@admin_bp.route("/logs/exportar")
@admin_required
@permiso_requerido(Permiso.EXPORTAR_DATOS)
def admin_logs_exportar():
    formato = request.args.get('formato', 'excel')
    
    filtros = {}
    if request.args.get('modulo'):
        filtros['modulo'] = request.args.get('modulo')
    if request.args.get('nivel'):
        filtros['nivel'] = request.args.get('nivel')
    if request.args.get('usuario'):
        filtros['usuario_email'] = request.args.get('usuario')
    if request.args.get('fecha_desde'):
        try:
            filtros['fecha_desde'] = datetime.strptime(request.args.get('fecha_desde'), '%Y-%m-%d')
        except:
            pass
    if request.args.get('fecha_hasta'):
        try:
            filtros['fecha_hasta'] = datetime.strptime(request.args.get('fecha_hasta'), '%Y-%m-%d')
        except:
            pass
    
    logs_data = LogActividad.exportar_a_lista(filtros=filtros, limite=5000)
    
    if formato == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Logs de Actividad"
            
            headers = ['Fecha', 'Usuario', 'Email', 'Acción', 'Módulo', 'Nivel', 'Descripción', 'IP']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, log in enumerate(logs_data, 2):
                ws.cell(row=row, column=1, value=log.get('fecha_formateada', ''))
                ws.cell(row=row, column=2, value=log.get('usuario_nombre', ''))
                ws.cell(row=row, column=3, value=log.get('usuario_email', ''))
                ws.cell(row=row, column=4, value=log.get('accion', ''))
                ws.cell(row=row, column=5, value=log.get('modulo', ''))
                ws.cell(row=row, column=6, value=log.get('nivel', ''))
                ws.cell(row=row, column=7, value=log.get('descripcion', '')[:200])
                ws.cell(row=row, column=8, value=log.get('ip_address', ''))
            
            for col in range(1, 9):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"logs_actividad_{fecha_str}.xlsx"
            
            registrar_log('exportar_logs', 'admin', f"Exportó logs a Excel", nivel='info')
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except ImportError:
            flash("❌ La exportación a Excel no está disponible. Instala openpyxl: pip install openpyxl", "error")
            return redirect(url_for('admin.admin_logs'))
    else:
        flash("❌ Formato no soportado", "error")
        return redirect(url_for('admin.admin_logs'))


# ================================================================
# BITÁCORA (LEGACY) (sin cambios)
# ================================================================

@admin_bp.route("/bitacora")
@admin_required
@permiso_requerido(Permiso.VER_BITACORA)
def bitacora():
    acciones_file = "data/admin_actions.json"
    acciones = []

    if os.path.exists(acciones_file):
        try:
            with open(acciones_file, 'r', encoding='utf-8') as f:
                acciones = json.load(f)
        except Exception as e:
            print(f"Error cargando bitácora: {e}")
            acciones = []

    acciones.sort(key=lambda x: x.get('fecha', ''), reverse=True)

    pagina = int(request.args.get('pagina', 1))
    por_pagina = 50
    inicio = (pagina - 1) * por_pagina
    fin = inicio + por_pagina
    total_pags = (len(acciones) + por_pagina - 1) // por_pagina if acciones else 1

    return render_template(
        "admin/bitacora.html",
        acciones=acciones[inicio:fin],
        pagina=pagina,
        total_paginas=total_pags,
        total_acciones=len(acciones)
    )


# ================================================================
# GESTIÓN DE MENSAJES EN TRÁMITES (sin cambios)
# ================================================================

from models.mensaje import Mensaje

@admin_bp.route("/api/tramite/<folio>/mensajes", methods=["GET"])
@admin_required
def api_obtener_mensajes_tramite(folio):
    try:
        tramite_tipo = request.args.get('tipo', 'solicitud')
        mensajes = Mensaje.query.filter_by(tramite_folio=folio, tramite_tipo=tramite_tipo).order_by(Mensaje.fecha_creacion.asc()).all()
        
        return jsonify({
            'success': True,
            'mensajes': [m.to_dict() for m in mensajes]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route("/api/tramite/<folio>/responder", methods=["POST"])
@admin_required
def api_responder_mensaje_admin(folio):
    try:
        data = request.get_json()
        mensaje_texto = data.get('mensaje', '').strip()
        tramite_tipo = data.get('tipo', 'solicitud')
        
        if not mensaje_texto:
            return jsonify({'success': False, 'error': 'El mensaje no puede estar vacío'}), 400
        
        if len(mensaje_texto) > 1000:
            return jsonify({'success': False, 'error': 'El mensaje es demasiado largo'}), 400
        
        usuario_email = None
        nombre_tramite = ''
        
        if tramite_tipo == 'solicitud':
            solicitud = Solicitud.query.filter_by(folio=folio).first()
            if solicitud:
                usuario_email = solicitud.usuario_email
                nombre_tramite = NOMBRES_SERVICIOS.get(str(solicitud.servicio_id), solicitud.servicio_id)
        elif tramite_tipo == 'denuncia':
            denuncia = Denuncia.query.filter_by(folio=folio).first()
            if denuncia:
                usuario_email = denuncia.usuario_email
                nombre_tramite = NOMBRES_DENUNCIAS.get(denuncia.tipo, denuncia.tipo)
        elif tramite_tipo == 'cita':
            from models.cita import Cita
            cita = Cita.query.filter_by(folio=folio).first()
            if cita:
                usuario_email = cita.usuario_email
                nombre_tramite = SERVICIOS_CITAS.get(cita.servicio, cita.servicio)
        
        if not usuario_email:
            return jsonify({'success': False, 'error': 'Trámite no encontrado'}), 404
        
        admin_email = session.get('user')
        admin_nombre = session.get('user_name', 'Administrador')
        
        Mensaje.crear_mensaje(
            tramite_folio=folio,
            tramite_tipo=tramite_tipo,
            usuario_email=usuario_email,
            autor_email=admin_email,
            autor_nombre=admin_nombre,
            mensaje=mensaje_texto,
            es_admin=True
        )
        
        try:
            titulo = f"Nuevo mensaje en tu {nombre_tramite}"
            descripcion = f"El administrador ha respondido a tu trámite {folio}"
            
            Notificacion.crear_notificacion(
                usuario_email=usuario_email,
                tipo='mensaje',
                titulo=titulo,
                mensaje=descripcion,
                datos_extra={
                    'folio': folio,
                    'tipo': tramite_tipo,
                    'url': url_for('mis_tramites')
                }
            )
        except:
            pass
        
        registrar_accion('responder_tramite', f"Admin respondió en {tramite_tipo} {folio}")
        
        return jsonify({'success': True, 'mensaje': 'Respuesta enviada correctamente'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ================================================================
# GESTIÓN DE CONTENIDO (PÁGINAS ESTÁTICAS) (sin cambios)
# ================================================================

@admin_bp.route("/contenidos")
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_contenidos():
    contenidos = Contenido.query.order_by(Contenido.id).all()
    return render_template("admin/contenidos.html", contenidos=contenidos)


@admin_bp.route("/contenidos/crear", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_contenido_crear():
    try:
        slug = request.form.get('slug', '').strip()
        titulo = request.form.get('titulo', '').strip()
        contenido = request.form.get('contenido', '')
        meta_descripcion = request.form.get('meta_descripcion', '')
        palabras_clave = request.form.get('palabras_clave', '')
        activo = request.form.get('activo') == 'on'
        
        if not slug or not titulo:
            flash("❌ Slug y título son obligatorios.", "error")
            return redirect(url_for('admin.admin_contenidos'))
        
        if Contenido.query.filter_by(slug=slug).first():
            flash(f"❌ Ya existe un contenido con el slug '{slug}'.", "error")
            return redirect(url_for('admin.admin_contenidos'))
        
        nuevo = Contenido(
            slug=slug,
            titulo=titulo,
            contenido=contenido,
            meta_descripcion=meta_descripcion,
            palabras_clave=palabras_clave,
            activo=activo
        )
        db.session.add(nuevo)
        db.session.commit()
        
        registrar_log(
            accion='crear_contenido',
            modulo='contenido',
            descripcion=f"Creó la página '{titulo}' (slug: {slug})"
        )
        
        flash(f"✅ Página '{titulo}' creada exitosamente.", "success")
    except Exception as e:
        flash(f"❌ Error al crear: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_contenidos'))


@admin_bp.route("/contenidos/<int:id>/editar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_contenido_editar(id):
    try:
        contenido_obj = Contenido.query.get_or_404(id)
        
        contenido_obj.slug = request.form.get('slug', '').strip()
        contenido_obj.titulo = request.form.get('titulo', '').strip()
        contenido_obj.contenido = request.form.get('contenido', '')
        contenido_obj.meta_descripcion = request.form.get('meta_descripcion', '')
        contenido_obj.palabras_clave = request.form.get('palabras_clave', '')
        contenido_obj.activo = request.form.get('activo') == 'on'
        
        db.session.commit()
        
        registrar_log(
            accion='editar_contenido',
            modulo='contenido',
            descripcion=f"Editó la página '{contenido_obj.titulo}' (ID: {id})"
        )
        
        flash(f"✅ Página '{contenido_obj.titulo}' actualizada correctamente.", "success")
    except Exception as e:
        flash(f"❌ Error al editar: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_contenidos'))


@admin_bp.route("/contenidos/<int:id>/eliminar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_contenido_eliminar(id):
    try:
        contenido_obj = Contenido.query.get_or_404(id)
        titulo = contenido_obj.titulo
        
        db.session.delete(contenido_obj)
        db.session.commit()
        
        registrar_log(
            accion='eliminar_contenido',
            modulo='contenido',
            descripcion=f"Eliminó la página '{titulo}' (ID: {id})"
        )
        
        flash(f"✅ Página '{titulo}' eliminada correctamente.", "success")
    except Exception as e:
        flash(f"❌ Error al eliminar: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_contenidos'))


# ================================================================
# GESTIÓN DE TRANSPARENCIA (DOCUMENTOS PÚBLICOS) (sin cambios)
# ================================================================

@admin_bp.route("/transparencia")
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_transparencia():
    documentos = Transparencia.query.order_by(Transparencia.id).all()
    return render_template("admin/transparencia.html", documentos=documentos, categorias=Transparencia.CATEGORIAS)


@admin_bp.route("/transparencia/crear", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_transparencia_crear():
    try:
        titulo = request.form.get('titulo', '').strip()
        descripcion = request.form.get('descripcion', '')
        categoria = request.form.get('categoria', 'general')
        anio = request.form.get('anio', type=int)
        periodo = request.form.get('periodo', '')
        archivo_url = request.form.get('archivo_url', '')
        enlace_ext = request.form.get('enlace_ext', '')
        publicado = request.form.get('publicado') == 'on'
        fecha_doc_str = request.form.get('fecha_doc', '')
        
        if not titulo:
            flash("❌ El título es obligatorio.", "error")
            return redirect(url_for('admin.admin_transparencia'))
        
        fecha_doc = None
        if fecha_doc_str:
            try:
                fecha_doc = datetime.strptime(fecha_doc_str, '%Y-%m-%d')
            except:
                pass
        
        nuevo = Transparencia(
            titulo=titulo,
            descripcion=descripcion,
            categoria=categoria,
            anio=anio,
            periodo=periodo,
            archivo_url=archivo_url,
            enlace_ext=enlace_ext,
            publicado=publicado,
            fecha_doc=fecha_doc
        )
        db.session.add(nuevo)
        db.session.commit()
        
        registrar_log(
            accion='crear_documento_transparencia',
            modulo='transparencia',
            descripcion=f"Creó documento de transparencia '{titulo}'"
        )
        
        flash(f"✅ Documento '{titulo}' creado exitosamente.", "success")
    except Exception as e:
        flash(f"❌ Error al crear: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_transparencia'))


@admin_bp.route("/transparencia/<int:id>/editar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_transparencia_editar(id):
    try:
        documento = Transparencia.query.get_or_404(id)
        
        documento.titulo = request.form.get('titulo', '').strip()
        documento.descripcion = request.form.get('descripcion', '')
        documento.categoria = request.form.get('categoria', 'general')
        documento.anio = request.form.get('anio', type=int) or None
        documento.periodo = request.form.get('periodo', '')
        documento.archivo_url = request.form.get('archivo_url', '')
        documento.enlace_ext = request.form.get('enlace_ext', '')
        documento.publicado = request.form.get('publicado') == 'on'
        
        fecha_doc_str = request.form.get('fecha_doc', '')
        if fecha_doc_str:
            try:
                documento.fecha_doc = datetime.strptime(fecha_doc_str, '%Y-%m-%d')
            except:
                pass
        
        db.session.commit()
        
        registrar_log(
            accion='editar_documento_transparencia',
            modulo='transparencia',
            descripcion=f"Editó documento '{documento.titulo}'"
        )
        
        flash(f"✅ Documento '{documento.titulo}' actualizado correctamente.", "success")
    except Exception as e:
        flash(f"❌ Error al editar: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_transparencia'))


@admin_bp.route("/transparencia/<int:id>/toggle", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_transparencia_toggle(id):
    try:
        documento = Transparencia.query.get_or_404(id)
        documento.publicado = not documento.publicado
        db.session.commit()
        
        estado = "publicado" if documento.publicado else "archivado"
        registrar_log(
            accion='toggle_transparencia',
            modulo='transparencia',
            descripcion=f"Cambió estado de '{documento.titulo}' a {estado}"
        )
        
        flash(f"✅ Documento '{documento.titulo}' {estado}.", "success")
    except Exception as e:
        flash(f"❌ Error: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_transparencia'))


@admin_bp.route("/transparencia/<int:id>/eliminar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_transparencia_eliminar(id):
    try:
        documento = Transparencia.query.get_or_404(id)
        titulo = documento.titulo
        
        db.session.delete(documento)
        db.session.commit()
        
        registrar_log(
            accion='eliminar_documento_transparencia',
            modulo='transparencia',
            descripcion=f"Eliminó documento de transparencia '{titulo}'"
        )
        
        flash(f"✅ Documento '{titulo}' eliminado correctamente.", "success")
    except Exception as e:
        flash(f"❌ Error al eliminar: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_transparencia'))


# ================================================================
# GESTIÓN DE MENÚ (NAVEGACIÓN DINÁMICA) (sin cambios)
# ================================================================

@admin_bp.route("/menu")
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_menu():
    items = MenuItem.query.order_by(MenuItem.orden).all()
    return render_template("admin/menu.html", items=items)


@admin_bp.route("/menu/crear", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_menu_crear():
    try:
        titulo = request.form.get('titulo', '').strip()
        url = request.form.get('url', '').strip()
        icono = request.form.get('icono', '')
        parent_id = request.form.get('parent_id', type=int) or None
        orden = request.form.get('orden', 0, type=int)
        activo = request.form.get('activo') == 'on'
        target_blank = request.form.get('target_blank') == 'on'
        requiere_login = request.form.get('requiere_login') == 'on'
        roles_permitidos = request.form.get('roles_permitidos', '')
        
        if not titulo:
            flash("❌ El título es obligatorio.", "error")
            return redirect(url_for('admin.admin_menu'))
        
        if not url:
            url = '#'
        
        nuevo = MenuItem(
            titulo=titulo,
            url=url,
            icono=icono,
            parent_id=parent_id,
            orden=orden,
            activo=activo,
            target_blank=target_blank,
            requiere_login=requiere_login,
            roles_permitidos=roles_permitidos
        )
        db.session.add(nuevo)
        db.session.commit()
        
        registrar_log(
            accion='crear_menu_item',
            modulo='menu',
            descripcion=f"Creó ítem de menú '{titulo}'"
        )
        
        flash(f"✅ Ítem de menú '{titulo}' creado exitosamente.", "success")
    except Exception as e:
        flash(f"❌ Error al crear: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_menu'))


@admin_bp.route("/menu/<int:id>/editar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_menu_editar(id):
    try:
        item = MenuItem.query.get_or_404(id)
        
        item.titulo = request.form.get('titulo', '').strip()
        item.url = request.form.get('url', '').strip() or '#'
        item.icono = request.form.get('icono', '')
        item.parent_id = request.form.get('parent_id', type=int) or None
        item.orden = request.form.get('orden', 0, type=int)
        item.activo = request.form.get('activo') == 'on'
        item.target_blank = request.form.get('target_blank') == 'on'
        item.requiere_login = request.form.get('requiere_login') == 'on'
        item.roles_permitidos = request.form.get('roles_permitidos', '')
        
        db.session.commit()
        
        registrar_log(
            accion='editar_menu_item',
            modulo='menu',
            descripcion=f"Editó ítem de menú '{item.titulo}'"
        )
        
        flash(f"✅ Ítem de menú '{item.titulo}' actualizado correctamente.", "success")
    except Exception as e:
        flash(f"❌ Error al editar: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_menu'))


@admin_bp.route("/menu/<int:id>/toggle", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_menu_toggle(id):
    try:
        item = MenuItem.query.get_or_404(id)
        item.activo = not item.activo
        db.session.commit()
        
        estado = "activado" if item.activo else "desactivado"
        registrar_log(
            accion='toggle_menu_item',
            modulo='menu',
            descripcion=f"Cambió estado del ítem '{item.titulo}' a {estado}"
        )
        
        flash(f"✅ Ítem de menú '{item.titulo}' {estado}.", "success")
    except Exception as e:
        flash(f"❌ Error: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_menu'))


@admin_bp.route("/menu/<int:id>/eliminar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_menu_eliminar(id):
    try:
        item = MenuItem.query.get_or_404(id)
        titulo = item.titulo
        
        hijos = MenuItem.query.filter_by(parent_id=id).count()
        if hijos > 0:
            flash(f"❌ No se puede eliminar '{titulo}' porque tiene {hijos} subítems asociados.", "error")
            return redirect(url_for('admin.admin_menu'))
        
        db.session.delete(item)
        db.session.commit()
        
        registrar_log(
            accion='eliminar_menu_item',
            modulo='menu',
            descripcion=f"Eliminó ítem de menú '{titulo}'"
        )
        
        flash(f"✅ Ítem de menú '{titulo}' eliminado correctamente.", "success")
    except Exception as e:
        flash(f"❌ Error al eliminar: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_menu'))


@admin_bp.route("/menu/reordenar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def admin_menu_reordenar():
    try:
        data = request.get_json()
        
        if not data or 'items' not in data:
            return jsonify({"success": False, "error": "Datos inválidos"}), 400
        
        items_data = data.get('items', [])
        
        for item_info in items_data:
            item_id = item_info.get('id')
            nuevo_orden = item_info.get('orden')
            nuevo_parent = item_info.get('parent_id', None)
            
            if item_id is not None and nuevo_orden is not None:
                item = MenuItem.query.get(item_id)
                if item:
                    item.orden = nuevo_orden
                    if nuevo_parent is not None:
                        item.parent_id = nuevo_parent if nuevo_parent else None
        
        db.session.commit()
        
        registrar_log(
            accion='reordenar_menu',
            modulo='menu',
            descripcion=f"Reordenó {len(items_data)} ítems del menú"
        )
        
        return jsonify({"success": True, "message": "Orden del menú actualizado correctamente"})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al reordenar menú: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route("/api/menu-item/<int:id>")
@admin_required
def api_menu_item(id):
    try:
        item = MenuItem.query.get_or_404(id)
        return jsonify({
            'success': True,
            'item': {
                'id': item.id,
                'titulo': item.titulo,
                'url': item.url,
                'icono': item.icono,
                'parent_id': item.parent_id,
                'orden': item.orden,
                'activo': item.activo,
                'target_blank': item.target_blank,
                'requiere_login': item.requiere_login,
                'roles_permitidos': item.roles_permitidos
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ================================================================
# API PARA TRANSPARENCIA (OBTENER DOCUMENTO) (sin cambios)
# ================================================================

@admin_bp.route("/api/transparencia/<int:id>")
@admin_required
def api_transparencia_item(id):
    try:
        doc = Transparencia.query.get_or_404(id)
        return jsonify({
            'success': True,
            'documento': {
                'id': doc.id,
                'titulo': doc.titulo,
                'descripcion': doc.descripcion or '',
                'categoria': doc.categoria,
                'anio': doc.anio,
                'periodo': doc.periodo or '',
                'archivo_url': doc.archivo_url or '',
                'enlace_ext': doc.enlace_ext or '',
                'publicado': doc.publicado,
                'fecha_doc': doc.fecha_doc.isoformat() if doc.fecha_doc else None
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ================================================================
# CONFIGURACIÓN (sin cambios)
# ================================================================

@admin_bp.route("/configuracion", methods=["GET"])
@admin_required
@permiso_requerido(Permiso.VER_CONFIG)
def configuracion():
    try:
        from models.configuracion import Configuracion
        registros = Configuracion.query.all()
        config_actual = {r.clave: r.get_valor() for r in registros}
    except Exception as e:
        print(f"Error cargando config: {e}")
        config_actual = {}
    return render_template("admin/configuracion.html", config=config_actual)


@admin_bp.route("/configuracion/guardar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def guardar_configuracion():
    from models.configuracion import Configuracion
    
    for clave, valor in request.form.items():
        Configuracion.set(clave, valor)
    
    Configuracion.clear_cache()
    
    flash('✅ Configuración actualizada correctamente', 'success')
    return redirect(url_for('admin.configuracion'))


# ================================================================
# API DE CONFIGURACIÓN (sin cambios)
# ================================================================

@admin_bp.route("/api/config/sistema-info", methods=["GET"])
@admin_required
@permiso_requerido(Permiso.VER_CONFIG)
def api_sistema_info():
    try:
        info = {
            'version': '2.0.0',
            'python_version': sys.version.split()[0],
            'flask_version': flask.__version__,
            'database': 'PostgreSQL',
            'debug': True,
            'redis_available': REDIS_AVAILABLE,
            'cloudinary_configured': True,
            'total_usuarios': Usuario.query.count(),
            'total_solicitudes': Solicitud.query.count(),
            'total_denuncias': Denuncia.query.count(),
            'ultimo_backup': None,
            'mantenimiento': cfg.get('sistema', 'maintenance', False)
        }
        return jsonify({"ok": True, "datos": info})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@admin_bp.route("/api/config/guardar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def api_guardar_config():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"ok": False, "error": "No se recibieron datos"}), 400
        
        seccion = data.get('seccion')
        valores = data.get('datos', {})
        
        if not valores:
            return jsonify({"ok": False, "error": "No se recibieron valores para guardar"}), 400
        
        from models.configuracion import Configuracion
        
        contador = 0
        if seccion:
            for clave, valor in valores.items():
                tipo = 'string'
                if isinstance(valor, bool):
                    tipo = 'bool'
                elif isinstance(valor, int):
                    tipo = 'int'
                elif isinstance(valor, float):
                    tipo = 'float'
                elif isinstance(valor, (dict, list)):
                    tipo = 'json'
                
                Configuracion.set(clave, valor, tipo, seccion)
                contador += 1
        else:
            for clave, valor in valores.items():
                tipo = 'string'
                if isinstance(valor, bool):
                    tipo = 'bool'
                elif isinstance(valor, int):
                    tipo = 'int'
                elif isinstance(valor, float):
                    tipo = 'float'
                elif isinstance(valor, (dict, list)):
                    tipo = 'json'
                Configuracion.set(clave, valor, tipo, 'general')
                contador += 1
        
        db.session.commit()
        
        registrar_accion(
            'guardar_configuracion', 
            f"Configuración actualizada - Sección: {seccion or 'general'} - {contador} valores guardados"
        )
        
        return jsonify({
            "ok": True, 
            "mensaje": f"✅ Configuración guardada correctamente ({contador} valores actualizados)",
            "guardados": contador
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error guardando configuración: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@admin_bp.route("/api/config/test-smtp", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def api_test_smtp():
    try:
        data = request.get_json()
        exito, mensaje = cfg.test_smtp(
            host=data.get('smtp_host', ''),
            port=data.get('smtp_port', 587),
            user=data.get('smtp_user', ''),
            password=data.get('smtp_pass', ''),
            nombre=data.get('smtp_name', 'Villa Cutupú'),
            email_destino=data.get('email_destino', '')
        )
        return jsonify({"ok": exito, "msg": mensaje})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@admin_bp.route("/api/config/subir-imagen", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_CONFIG)
def api_subir_imagen():
    try:
        tipo = request.form.get('tipo')
        archivo = request.files.get('archivo')
        
        if not tipo or not archivo or archivo.filename == '':
            return jsonify({"ok": False, "msg": "Tipo y archivo requeridos"}), 400
        
        exito, resultado = cfg.guardar_imagen_config(archivo, tipo)
        
        if exito:
            registrar_accion('subir_imagen', f"Imagen '{tipo}' actualizada")
            return jsonify({"ok": True, "ruta": resultado, "msg": f"{tipo.capitalize()} guardado correctamente"})
        else:
            return jsonify({"ok": False, "msg": resultado}), 400
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@admin_bp.route("/api/config/exportar", methods=["GET"])
@admin_required
@permiso_requerido(Permiso.EXPORTAR_DATOS)
def api_exportar_datos():
    try:
        exito, ruta = cfg.exportar_datos_zip()
        if exito:
            registrar_accion('exportar_datos', "Exportación ZIP generada")
            return send_file(
                ruta,
                as_attachment=True,
                download_name=Path(ruta).name,
                mimetype='application/zip'
            )
        else:
            return jsonify({"ok": False, "msg": ruta}), 500
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@admin_bp.route("/api/config/limpiar-cache", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.MANTENIMIENTO)
def api_limpiar_cache():
    try:
        if cache:
            cache.clear()
        registrar_accion('limpiar_cache', "Caché del sistema limpiada")
        return jsonify({"ok": True, "msg": "✅ Caché limpiada correctamente"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@admin_bp.route("/api/config/mantenimiento", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.MANTENIMIENTO)
def api_mantenimiento():
    try:
        data = request.get_json()
        activo = data.get('activo', False)
        cfg.guardar_seccion('sistema', {'maintenance': activo})
        estado = 'activado' if activo else 'desactivado'
        registrar_accion('mantenimiento', f"Modo mantenimiento {estado}")
        return jsonify({"ok": True, "msg": f"Modo mantenimiento {estado}"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# ================================================================
# API PARA ADMIN (ENDPOINTS SIN @admin_required PARA AJAX) (sin cambios)
# ================================================================

@admin_bp.route("/api/citas-pendientes")
def api_citas_pendientes():
    if "user" not in session:
        return jsonify({"count": 0}), 401
    try:
        from models.cita import Cita
        pendientes = Cita.query.filter_by(estado='pendiente').count()
        return jsonify({"count": pendientes})
    except Exception as e:
        return jsonify({"count": 0, "error": str(e)}), 500


@admin_bp.route("/api/solicitudes-pendientes")
def api_solicitudes_pendientes():
    if "user" not in session:
        return jsonify({"count": 0}), 401
    try:
        pendientes = Solicitud.query.filter(Solicitud.estado.in_(['pendiente', 'en_proceso'])).count()
        return jsonify({"count": pendientes})
    except Exception as e:
        return jsonify({"count": 0, "error": str(e)}), 500


@admin_bp.route("/api/denuncias-pendientes")
def api_denuncias_pendientes():
    if "user" not in session:
        return jsonify({"count": 0}), 401
    try:
        pendientes = Denuncia.query.filter(Denuncia.estado.in_(['pendiente', 'en_investigacion'])).count()
        return jsonify({"count": pendientes})
    except Exception as e:
        return jsonify({"count": 0, "error": str(e)}), 500


# ================================================================
# API PARA DENUNCIAS EN MAPA (GEOJSON) (sin cambios)
# ================================================================

@admin_bp.route("/api/denuncias/geojson")
@admin_required
def api_denuncias_geojson():
    try:
        denuncias = Denuncia.query.all()
        
        features = []
        for d in denuncias:
            lat = None
            lng = None
            
            if hasattr(d, 'lat') and d.lat and hasattr(d, 'lng') and d.lng:
                try:
                    lat = float(d.lat)
                    lng = float(d.lng)
                except (ValueError, TypeError):
                    pass
            elif hasattr(d, 'latitud') and d.latitud and hasattr(d, 'longitud') and d.longitud:
                try:
                    lat = float(d.latitud)
                    lng = float(d.longitud)
                except (ValueError, TypeError):
                    pass
            
            if lat and lng:
                tipo_nombre = NOMBRES_DENUNCIAS.get(d.tipo, d.tipo)
                
                color = "#ffc107"
                if d.estado == 'pendiente':
                    color = "#ffc107"
                elif d.estado == 'en_investigacion':
                    color = "#17a2b8"
                elif d.estado == 'resuelto':
                    color = "#28a745"
                elif d.estado == 'rechazado':
                    color = "#dc3545"
                
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [lng, lat]
                    },
                    "properties": {
                        "id": d.id,
                        "folio": d.folio,
                        "tipo": d.tipo,
                        "tipo_nombre": tipo_nombre,
                        "estado": d.estado,
                        "estado_label": d.estado.replace('_', ' ').title(),
                        "color": color,
                        "descripcion": d.descripcion[:200] if d.descripcion else "",
                        "direccion": d.direccion or "",
                        "fecha": d.fecha_creacion.isoformat() if d.fecha_creacion else None,
                        "url": url_for('admin.detalle_denuncia', denuncia_id=d.id)
                    }
                }
                features.append(feature)
        
        geojson = {
            "type": "FeatureCollection",
            "features": features,
            "total": len(features),
            "total_denuncias": len(denuncias)
        }
        
        return jsonify(geojson)
        
    except Exception as e:
        print(f"Error generando GeoJSON para mapa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "type": "FeatureCollection",
            "features": [],
            "error": str(e),
            "total": 0
        }), 500


@admin_bp.route("/api/estadisticas")
@admin_required
@permiso_requerido(Permiso.VER_BITACORA)
def api_estadisticas():
    try:
        stats = Reportes.obtener_estadisticas_generales()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ================================================================
# API PARA CONTACTOS PENDIENTES (sin cambios)
# ================================================================

@admin_bp.route("/api/contactos-pendientes")
def api_contactos_pendientes():
    if "user" not in session:
        return jsonify({"count": 0}), 401
    try:
        from models.mensaje import Mensaje
        pendientes = Mensaje.query.filter_by(tramite_tipo='consulta', es_admin=False).count()
        return jsonify({"count": pendientes})
    except Exception as e:
        return jsonify({"count": 0, "error": str(e)}), 500


# ================================================================
# RUTA PARA GESTIÓN DE CONTACTOS (ADMIN) (sin cambios)
# ================================================================

@admin_bp.route("/contactos")
@admin_required
@permiso_requerido(Permiso.VER_SOLICITUDES)
def admin_contactos():
    try:
        from models.mensaje import Mensaje
        contactos = Mensaje.obtener_todos_contactos()
        return render_template("admin/contactos.html", contactos=contactos)
    except Exception as e:
        flash(f"Error al cargar contactos: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


@admin_bp.route("/contactos/<int:contacto_id>/responder", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.EDITAR_SOLICITUDES)
def admin_contacto_responder(contacto_id):
    try:
        from models.mensaje import Mensaje
        respuesta = request.form.get('respuesta', '').strip()
        
        if not respuesta:
            flash("❌ La respuesta no puede estar vacía.", "error")
            return redirect(url_for('admin.admin_contactos'))
        
        admin_email = session.get('user')
        admin_nombre = session.get('user_name', 'Administrador')
        
        Mensaje.responder_contacto(contacto_id, admin_email, admin_nombre, respuesta)
        
        flash("✅ Respuesta enviada correctamente.", "success")
    except Exception as e:
        flash(f"❌ Error al enviar respuesta: {str(e)}", "error")
    
    return redirect(url_for('admin.admin_contactos'))


# ================================================================
# REPORTES AVANZADOS CON PERSONALIZACIÓN VISUAL (con logo por defecto)
# ================================================================

from utils.reportes_utils import (
    dataframe_desde_solicitudes, dataframe_desde_denuncias,
    dataframe_desde_usuarios, dataframe_desde_citas,
    dataframe_desde_contactos, generar_grafico_barras,
    generar_tabla_html_profesional, exportar_excel_profesional,
    generar_pdf_desde_html
)

@admin_bp.route("/reportes/crear", methods=["GET", "POST"])
@admin_required
@permiso_requerido(Permiso.VER_BITACORA)
def crear_reporte():
    if request.method == "POST":
        # ========== PARÁMETROS BÁSICOS ==========
        tipo = request.form.get("tipo_reporte")
        fecha_desde = request.form.get("fecha_desde")
        fecha_hasta = request.form.get("fecha_hasta")
        estados = request.form.getlist("estados")
        servicios_tipos = request.form.getlist("servicios_tipos")
        columnas_seleccionadas = request.form.getlist("columnas")
        formato = request.form.get("formato", "html")
        
        # ========== ESTILOS PERSONALIZADOS (capturados desde el formulario) ==========
        estilos = {
            'titulo_personalizado': request.form.get('titulo_personalizado', ''),
            'color_encabezado': request.form.get('color_encabezado', '#2D5016'),
            'color_texto_encabezado': request.form.get('color_texto_encabezado', '#FFFFFF'),
            'color_fila_par': request.form.get('color_fila_par', '#F8F9FA'),
            'color_fila_impar': request.form.get('color_fila_impar', '#FFFFFF'),
            'fuente': request.form.get('fuente', 'Arial, sans-serif'),
            'tamano_fuente': int(request.form.get('tamano_fuente', 14)),
            'formato_fecha': request.form.get('formato_fecha', '%d/%m/%Y %H:%M'),
            'logo_url': request.form.get('logo_url', ''),
            'pie_pagina': request.form.get('pie_pagina', ''),
            'mostrar_bordes': request.form.get('mostrar_bordes') == 'on',
            'mostrar_logo': request.form.get('mostrar_logo') == 'on'
        }
        if not estilos['titulo_personalizado']:
            estilos['titulo_personalizado'] = f"Reporte de {tipo.replace('_', ' ').title()}"
        
        # ========== CONSTRUIR FILTROS ==========
        filtros = {}
        if fecha_desde:
            filtros['fecha_desde'] = datetime.strptime(fecha_desde, '%Y-%m-%d')
        if fecha_hasta:
            filtros['fecha_hasta'] = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        if estados:
            filtros['estados'] = estados
        if servicios_tipos:
            filtros['servicios_tipos'] = servicios_tipos
        
        # ========== OBTENER DATOS SEGÚN TIPO ==========
        datos = None
        df = None
        if tipo == 'solicitudes':
            query = Solicitud.query
            if fecha_desde:
                query = query.filter(Solicitud.fecha_creacion >= filtros['fecha_desde'])
            if fecha_hasta:
                query = query.filter(Solicitud.fecha_creacion <= filtros['fecha_hasta'])
            if estados:
                query = query.filter(Solicitud.estado.in_(estados))
            if servicios_tipos:
                query = query.filter(Solicitud.servicio_id.in_(servicios_tipos))
            datos = query.all()
            df = dataframe_desde_solicitudes(datos, columnas_seleccionadas)
        elif tipo == 'denuncias':
            query = Denuncia.query
            if fecha_desde:
                query = query.filter(Denuncia.fecha_creacion >= filtros['fecha_desde'])
            if fecha_hasta:
                query = query.filter(Denuncia.fecha_creacion <= filtros['fecha_hasta'])
            if estados:
                query = query.filter(Denuncia.estado.in_(estados))
            if servicios_tipos:
                query = query.filter(Denuncia.tipo.in_(servicios_tipos))
            datos = query.all()
            df = dataframe_desde_denuncias(datos, columnas_seleccionadas)
        elif tipo == 'usuarios':
            query = Usuario.query
            if fecha_desde:
                query = query.filter(Usuario.fecha_registro >= filtros['fecha_desde'])
            if fecha_hasta:
                query = query.filter(Usuario.fecha_registro <= filtros['fecha_hasta'])
            if estados:
                if 'activo' in estados and 'inactivo' not in estados:
                    query = query.filter(Usuario.activo == True)
                elif 'inactivo' in estados and 'activo' not in estados:
                    query = query.filter(Usuario.activo == False)
            if servicios_tipos:
                query = query.filter(Usuario.tipo.in_(servicios_tipos))
            datos = query.all()
            df = dataframe_desde_usuarios(datos, columnas_seleccionadas)
        elif tipo == 'citas':
            from models.cita import Cita
            query = Cita.query
            if fecha_desde:
                query = query.filter(Cita.fecha_creacion >= filtros['fecha_desde'])
            if fecha_hasta:
                query = query.filter(Cita.fecha_creacion <= filtros['fecha_hasta'])
            if estados:
                query = query.filter(Cita.estado.in_(estados))
            if servicios_tipos:
                query = query.filter(Cita.servicio.in_(servicios_tipos))
            datos = query.all()
            df = dataframe_desde_citas(datos, columnas_seleccionadas)
        elif tipo == 'contactos':
            from models.mensaje import Mensaje
            query = Mensaje.query.filter_by(tramite_tipo='consulta', es_admin=False)
            if fecha_desde:
                query = query.filter(Mensaje.fecha_creacion >= filtros['fecha_desde'])
            if fecha_hasta:
                query = query.filter(Mensaje.fecha_creacion <= filtros['fecha_hasta'])
            if estados:
                if 'respondido' in estados:
                    query = query.filter(Mensaje.tramite_folio.in_(
                        db.session.query(Mensaje.tramite_folio).filter(Mensaje.es_admin==True)
                    ))
                elif 'pendiente' in estados:
                    query = query.filter(~Mensaje.tramite_folio.in_(
                        db.session.query(Mensaje.tramite_folio).filter(Mensaje.es_admin==True)
                    ))
            datos = query.all()
            df = dataframe_desde_contactos(datos, columnas_seleccionadas)
        
        if df is None or df.empty:
            flash("No hay datos con los filtros seleccionados.", "warning")
            return redirect(url_for('admin.crear_reporte'))
        
        # ========== PARA VISTA PREVIA HTML ==========
        if formato == 'html':
            # Guardar todo en sesión temporal para la vista
            session['reporte_temporal'] = {
                'tipo': tipo,
                'filtros': filtros,
                'columnas': columnas_seleccionadas,
                'estilos': estilos,
                'datos_json': df.to_dict(orient='records')
            }
            return redirect(url_for('admin.mostrar_reporte_generado'))
        
        # ========== EXPORTAR A EXCEL ==========
        elif formato == 'excel':
            titulo_excel = estilos['titulo_personalizado'][:31]
            excel_file = exportar_excel_profesional(df, titulo_excel, estilos)
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
        # ========== EXPORTAR A PDF ==========
        elif formato == 'pdf':
            tabla_html = generar_tabla_html_profesional(df, estilos['titulo_personalizado'], estilos)
            filtros_serializables = {
                k: str(v) if not isinstance(v, (str, list, dict, int, float, bool, type(None))) else v
                for k, v in filtros.items()
            }
            context = {
                'titulo':                 estilos['titulo_personalizado'],
                'fecha_generacion':       datetime.now(),
                'total_registros':        len(datos),
                'tabla_html':             tabla_html,
                'estilos':                estilos,
                'nombre_municipio':       cfg.get('general', 'nombre_municipio', 'Villa Cutupú'),
                'usuario_genero':         session.get('user_name', 'Administrador'),
                'logo_url':               estilos.get('logo_url') or url_for('static', filename='img/logo.png'),
                'tipo_reporte':           tipo,
                'filtros_usados':         filtros_serializables,
                'columnas_seleccionadas': columnas_seleccionadas or [],
                'datos_json':             json.dumps(df.to_dict(orient='records'), default=str),
                'grafico_base64':         None,
                'now':                    datetime.now(),
            }
            pdf_io = generar_pdf_desde_template('admin/reporte_pdf.html', context)
            return send_file(
                pdf_io,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f"{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
    
    # GET: mostrar formulario con listas desplegables
    servicios_disponibles = [{'id': k, 'nombre': v} for k, v in NOMBRES_SERVICIOS.items()]
    tipos_denuncia = [{'id': k, 'nombre': v} for k, v in NOMBRES_DENUNCIAS.items()]
    estados_solicitud = ['pendiente', 'en_proceso', 'completado', 'cancelado']
    estados_denuncia = ['pendiente', 'en_investigacion', 'resuelto', 'rechazado']
    tipos_usuario = ['ciudadano', 'admin']
    servicios_citas = [{'id': k, 'nombre': v} for k, v in SERVICIOS_CITAS.items()]
    estados_citas = ['pendiente', 'confirmada', 'cancelada', 'completada']
    
    plantillas = []
    if ReportePlantilla is not None:
        try:
            plantillas = ReportePlantilla.query.filter_by(created_by=session.get('user')).all()
        except:
            plantillas = []
    
    return render_template("admin/crear_reporte.html",
                         servicios=servicios_disponibles,
                         tipos_denuncia=tipos_denuncia,
                         estados_solicitud=estados_solicitud,
                         estados_denuncia=estados_denuncia,
                         tipos_usuario=tipos_usuario,
                         servicios_citas=servicios_citas,
                         estados_citas=estados_citas,
                         plantillas=plantillas)


@admin_bp.route("/reportes/mostrar")
@admin_required
def mostrar_reporte_generado():
    """Vista previa del reporte generado, con opción de guardar"""
    reporte_temp = session.get('reporte_temporal')
    if not reporte_temp:
        flash("No hay reporte generado. Por favor, genere uno primero.", "warning")
        return redirect(url_for('admin.crear_reporte'))
    
    tipo = reporte_temp['tipo']
    estilos = reporte_temp['estilos']
    columnas = reporte_temp['columnas']
    datos = reporte_temp['datos_json']
    filtros = reporte_temp.get('filtros', {})
    
    df = pd.DataFrame(datos)
    if columnas:
        df = df[[c for c in columnas if c in df.columns]]
    
    tabla_html = generar_tabla_html_profesional(df, estilos['titulo_personalizado'], estilos)
    
    logo_url = estilos.get('logo_url', '') or url_for('static', filename='img/logo.png')
    
    return render_template("admin/reporte_generado.html",
                         titulo=estilos['titulo_personalizado'],
                         fecha_generacion=datetime.now(),
                         total_registros=len(datos),
                         tabla_html=tabla_html,
                         estilos=estilos,
                         tipo_reporte=tipo,
                         datos_json=json.dumps(datos, default=str),
                         columnas_seleccionadas=columnas,
                         filtros_usados=filtros,
                         logo_url=logo_url,
                         nombre_municipio=cfg.get('general', 'nombre_municipio', 'Villa Cutupú'),
                         usuario_genero=session.get('user_name', 'Administrador'),
                         now=datetime.now())


# ================================================================
# RUTA CORREGIDA: HISTORIAL DE REPORTES
# ================================================================

@admin_bp.route("/reportes/historial")
@admin_required
@permiso_requerido(Permiso.VER_BITACORA)
def reportes_historial():
    """Muestra todos los reportes guardados por el administrador"""
    # Obtener todos los reportes ordenados por fecha descendente
    reportes = ReporteGuardado.query.order_by(
        ReporteGuardado.fecha_generacion.desc()
    ).all()
    
    return render_template(
        "admin/reportes_guardados.html",
        reportes=reportes,
        nombre_municipio=cfg.get('general', 'nombre_municipio', 'Villa Cutupú'),
        now=datetime.now()
    )


# ================================================================
# FUNCIÓN CORREGIDA: GUARDAR REPORTE (USA EL NOMBRE EDITADO)
# ================================================================

@admin_bp.route("/reportes/guardar", methods=["POST"])
@admin_required
def guardar_reporte():
    """Guarda un reporte generado (incluyendo estilos) en la base de datos"""
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        tipo_reporte = data.get('tipo_reporte')
        datos = data.get('datos', [])
        columnas = data.get('columnas', [])
        filtros = data.get('filtros', {})
        estilos = data.get('estilos', {})
        
        if not nombre:
            return jsonify({'success': False, 'error': 'El nombre del reporte es obligatorio'}), 400
        
        # ✅ Usar el nombre editado desde el frontend (estilos.generado_por)
        # Si viene vacío, se usa el email de sesión o 'Administrador' como respaldo
        autor = estilos.get('generado_por', '').strip()
        if not autor:
            autor = session.get('user', 'Administrador')
        
        reporte = ReporteGuardado(
            nombre=nombre,
            tipo_reporte=tipo_reporte,
            generado_por=autor      # ← Ahora guarda el nombre editado
        )
        reporte.set_datos(datos)
        reporte.set_columnas(columnas)
        reporte.set_filtros(filtros)
        reporte.set_estilos(estilos)
        
        db.session.add(reporte)
        db.session.commit()
        
        registrar_accion('guardar_reporte', f"Guardó reporte '{nombre}' con {len(datos)} registros")
        return jsonify({'success': True, 'id': reporte.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ================================================================
# OTRAS RUTAS DE REPORTES (ya existentes)
# ================================================================

@admin_bp.route("/reportes/guardados")
@admin_required
def reportes_guardados():
    """Lista todos los reportes guardados por el administrador (redirige al historial)"""
    return redirect(url_for('admin.reportes_historial'))


@admin_bp.route("/reportes/ver/<int:reporte_id>")
@admin_required
def ver_reporte_guardado(reporte_id):
    reporte = ReporteGuardado.query.get_or_404(reporte_id)
    datos = reporte.get_datos()
    columnas = reporte.get_columnas()
    estilos = reporte.get_estilos()
    
    if not estilos:
        estilos = {
            'titulo_personalizado': reporte.nombre,
            'color_encabezado': '#2D5016',
            'color_texto_encabezado': '#FFFFFF',
            'color_fila_par': '#F8F9FA',
            'color_fila_impar': '#FFFFFF',
            'fuente': 'Arial, sans-serif',
            'tamano_fuente': 14,
            'formato_fecha': '%d/%m/%Y %H:%M',
            'mostrar_bordes': True,
            'mostrar_logo': True,
            'logo_url': ''
        }
    
    if datos:
        df = pd.DataFrame(datos)
        if columnas:
            df = df[[c for c in columnas if c in df.columns]]
        tabla_html = generar_tabla_html_profesional(df, estilos.get('titulo_personalizado', reporte.nombre), estilos)
    else:
        tabla_html = "<div class='alert alert-warning'>El reporte no contiene datos.</div>"
    
    # Logo por defecto si no se guardó ninguno
    logo_url = estilos.get('logo_url', '') or url_for('static', filename='img/logo.png')
    
    return render_template("admin/reporte_generado.html",
                         titulo=estilos.get('titulo_personalizado', reporte.nombre),
                         fecha_generacion=reporte.fecha_generacion,
                         total_registros=reporte.total_registros,
                         tabla_html=tabla_html,
                         estilos=estilos,
                         tipo_reporte=reporte.tipo_reporte,
                         datos_json=json.dumps(datos, default=str),
                         columnas_seleccionadas=columnas,
                         filtros_usados=reporte.get_filtros(),
                         logo_url=logo_url,
                         nombre_municipio=cfg.get('general', 'nombre_municipio', 'Villa Cutupú'),
                         usuario_genero=reporte.generado_por,   # ← Aquí se muestra el nombre guardado
                         es_guardado=True)


@admin_bp.route("/reportes/eliminar/<int:reporte_id>", methods=["POST"])
@admin_required
def eliminar_reporte_guardado(reporte_id):
    """Elimina un reporte previamente guardado."""
    try:
        reporte = ReporteGuardado.query.get_or_404(reporte_id)
        db.session.delete(reporte)
        db.session.commit()
        registrar_accion('eliminar_reporte', f"Eliminó el reporte '{reporte.nombre}'")
        return jsonify({'success': True, 'message': 'Reporte eliminado correctamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route("/reportes/guardar-plantilla", methods=["POST"])
@admin_required
def guardar_plantilla_reporte():
    if ReportePlantilla is None:
        return jsonify({'success': False, 'error': 'Módulo de plantillas no disponible'}), 500
    data = request.get_json()
    nombre = data.get('nombre')
    tipo_reporte = data.get('tipo_reporte')
    filtros = data.get('filtros', {})
    columnas = data.get('columnas', [])
    
    if not nombre or not tipo_reporte:
        return jsonify({'success': False, 'error': 'Faltan datos'}), 400
    
    plantilla = ReportePlantilla(
        nombre=nombre,
        tipo_reporte=tipo_reporte,
        created_by=session.get('user')
    )
    plantilla.set_filtros(filtros)
    plantilla.set_columnas(columnas)
    db.session.add(plantilla)
    db.session.commit()
    
    return jsonify({'success': True, 'id': plantilla.id})


@admin_bp.route("/reportes/cargar-plantilla/<int:id>")
@admin_required
def cargar_plantilla_reporte(id):
    if ReportePlantilla is None:
        return jsonify({'error': 'Módulo de plantillas no disponible'}), 500
    plantilla = ReportePlantilla.query.get_or_404(id)
    if plantilla.created_by != session.get('user') and session.get('user_rol') != 'super_admin':
        return jsonify({'error': 'No autorizado'}), 403
    return jsonify({
        'success': True,
        'tipo_reporte': plantilla.tipo_reporte,
        'filtros': plantilla.get_filtros(),
        'columnas': plantilla.get_columnas()
    })


@admin_bp.route("/reportes/generar", methods=["POST"])
@admin_required
@permiso_requerido(Permiso.VER_BITACORA)
def generar_reporte():
    """Genera reporte con columnas seleccionadas y filtros (respuesta JSON para AJAX)"""
    # Esta ruta se usa cuando se llama desde el formulario con AJAX.
    # Como ahora usamos /reportes/crear con redirección a /mostrar, esta ruta puede mantenerse
    # para compatibilidad, pero ya no es necesaria. Se deja por si existe código JS que la use.
    tipo = request.form.get("tipo_reporte")
    fecha_desde = request.form.get("fecha_desde")
    fecha_hasta = request.form.get("fecha_hasta")
    estados = request.form.getlist("estados")
    servicios_tipos = request.form.getlist("servicios_tipos")
    columnas = request.form.getlist("columnas")
    
    # Simplemente redirigimos a crear_reporte con los datos para no duplicar lógica
    # (opcional: procesar y devolver JSON)
    return redirect(url_for('admin.crear_reporte'))


# ================================================================
# VISTA UNIFICADA: TODOS LOS TRÁMITES (SOLICITUDES + DENUNCIAS)
# ================================================================

@admin_bp.route("/todos-tramites")
@admin_required
def todos_tramites():
    """Muestra todas las solicitudes y denuncias en una sola tabla (sin filtros)."""
    try:
        solicitudes = Solicitud.query.all()
        denuncias = Denuncia.query.all()
        
        tramites = []
        
        for s in solicitudes:
            tramites.append({
                'tipo': 'Solicitud',
                'folio': s.folio,
                'nombre': NOMBRES_SERVICIOS.get(str(s.servicio_id), s.servicio_id),
                'usuario': s.usuario_nombre or s.usuario_email,
                'estado': s.estado,
                'fecha': s.fecha_creacion.strftime('%Y-%m-%d %H:%M') if s.fecha_creacion else '',
                'url': url_for('admin.detalle_solicitud', solicitud_id=s.id)
            })
        
        for d in denuncias:
            tramites.append({
                'tipo': 'Denuncia',
                'folio': d.folio,
                'nombre': NOMBRES_DENUNCIAS.get(d.tipo, d.tipo),
                'usuario': d.usuario_nombre or d.usuario_email,
                'estado': d.estado,
                'fecha': d.fecha_creacion.strftime('%Y-%m-%d %H:%M') if d.fecha_creacion else '',
                'url': url_for('admin.detalle_denuncia', denuncia_id=d.id)
            })
        
        tramites.sort(key=lambda x: x['fecha'], reverse=True)
        
        stats = {
            'total': len(tramites),
            'solicitudes': len(solicitudes),
            'denuncias': len(denuncias),
            'pendientes': sum(1 for t in tramites if t['estado'] in ['pendiente', 'en_proceso', 'en_investigacion'])
        }
        
        return render_template("admin/todos_tramites.html", tramites=tramites, stats=stats)
    except Exception as e:
        flash(f"Error al cargar trámites: {str(e)}", "error")
        return redirect(url_for('admin.dashboard'))


# ================================================================
# HELPER - Registrar acción en bitácora
# ================================================================

def registrar_accion(tipo: str, descripcion: str, admin: str = None):
    acciones_file = "data/admin_actions.json"
    acciones = []

    if os.path.exists(acciones_file):
        try:
            with open(acciones_file, 'r', encoding='utf-8') as f:
                acciones = json.load(f)
        except:
            acciones = []

    acciones.append({
        'fecha': datetime.now().isoformat(),
        'tipo': tipo,
        'descripcion': descripcion,
        'admin': admin or session.get('user', 'desconocido')
    })

    if len(acciones) > 1000:
        acciones = acciones[-1000:]

    os.makedirs('data', exist_ok=True)
    try:
        with open(acciones_file, 'w', encoding='utf-8') as f:
            json.dump(acciones, f, ensure_ascii=False, indent=2)
    except:
        pass


# ================================================================
# CONTEXTO PARA PLANTILLAS ADMIN (CON COMENTARIOS PENDIENTES)
# ================================================================

@admin_bp.context_processor
def inject_admin_variables():
    user_email = session.get('user', '')
    foto_perfil = ''
    user_name = session.get('user_name', 'Administrador')
    user_rol = session.get('user_rol', 'admin')

    if user_email:
        try:
            usuario = Usuario.query.filter_by(email=user_email).first()
            if usuario and usuario.foto_perfil:
                foto_perfil = usuario.foto_perfil
        except Exception as e:
            pass

    try:
        comentarios_pendientes = ComentarioNoticia.query.filter_by(aprobado=False).count()
    except:
        comentarios_pendientes = 0

    return dict(
        ahora=datetime.now(),
        NOMBRES_SERVICIOS=NOMBRES_SERVICIOS,
        NOMBRES_DENUNCIAS=NOMBRES_DENUNCIAS,
        foto_perfil=foto_perfil,
        user_name=user_name,
        user_rol=user_rol,
        user_email=user_email,
        comentarios_pendientes=comentarios_pendientes
    )