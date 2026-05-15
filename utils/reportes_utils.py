# utils/reportes_utils.py
"""
Utilidades para la generación de reportes avanzados.
Incluye conversión a DataFrames, tablas HTML con estilos personalizables,
gráficos, exportación a Excel con formato y generación de PDF.
"""

import io
import base64
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Para entornos sin GUI
import matplotlib.pyplot as plt
from datetime import datetime
from flask import url_for
from weasyprint import HTML  # Para generar PDFs
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================================================================
# FUNCIONES PARA CREAR DATAFRAMES DESDE MODELOS
# ================================================================

def dataframe_desde_solicitudes(solicitudes, columnas_seleccionadas=None):
    """
    Convierte una lista de objetos Solicitud en un DataFrame de pandas.
    
    Args:
        solicitudes: Lista de instancias de Solicitud.
        columnas_seleccionadas: Lista opcional de nombres de columnas a incluir.
    
    Returns:
        pd.DataFrame con los datos filtrados.
    """
    data = []
    for s in solicitudes:
        row = {
            'ID': s.id,
            'Folio': s.folio,
            'Usuario': s.usuario_nombre,
            'Email': s.usuario_email,
            'Servicio': s.servicio_nombre,
            'Descripción': s.descripcion[:100] if s.descripcion else '',
            'Estado': s.estado,
            'Fecha Creación': s.fecha_creacion.strftime('%Y-%m-%d %H:%M') if s.fecha_creacion else '',
            'Última Actualización': s.fecha_actualizacion.strftime('%Y-%m-%d %H:%M') if s.fecha_actualizacion else ''
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    if columnas_seleccionadas:
        disponibles = [col for col in columnas_seleccionadas if col in df.columns]
        if disponibles:
            return df[disponibles]
    return df


def dataframe_desde_denuncias(denuncias, columnas_seleccionadas=None):
    """
    Convierte una lista de objetos Denuncia en un DataFrame.
    """
    data = []
    for d in denuncias:
        row = {
            'ID': d.id,
            'Folio': d.folio,
            'Usuario': d.usuario_nombre,
            'Email': d.usuario_email,
            'Tipo': getattr(d, 'tipo_nombre', d.tipo),
            'Ubicación': d.ubicacion or '',
            'Descripción': d.descripcion[:100] if d.descripcion else '',
            'Estado': d.estado,
            'Fecha': d.fecha_creacion.strftime('%Y-%m-%d %H:%M') if d.fecha_creacion else '',
            'Geolocalizada': 'Sí' if getattr(d, 'geolocalizada', False) else 'No'
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    if columnas_seleccionadas:
        disponibles = [col for col in columnas_seleccionadas if col in df.columns]
        if disponibles:
            return df[disponibles]
    return df


def dataframe_desde_usuarios(usuarios, columnas_seleccionadas=None):
    """
    Convierte una lista de objetos Usuario en un DataFrame.
    """
    data = []
    for u in usuarios:
        row = {
            'Email': u.email,
            'Nombre Completo': u.nombre_completo or '',
            'Cédula': u.cedula or '',
            'Teléfono': u.telefono or '',
            'Tipo': u.tipo,
            'Rol': u.rol or 'Ninguno',
            'Activo': 'Sí' if u.activo else 'No',
            'Registro': u.fecha_registro.strftime('%Y-%m-%d') if u.fecha_registro else ''
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    if columnas_seleccionadas:
        disponibles = [col for col in columnas_seleccionadas if col in df.columns]
        if disponibles:
            return df[disponibles]
    return df


def dataframe_desde_citas(citas, columnas_seleccionadas=None):
    """
    Convierte una lista de objetos Cita en un DataFrame.
    """
    data = []
    for c in citas:
        row = {
            'Folio': c.folio,
            'Usuario': c.usuario_nombre,
            'Email': c.usuario_email,
            'Servicio': c.servicio_nombre,
            'Fecha Cita': c.fecha,
            'Hora': c.hora,
            'Motivo': c.motivo[:100] if c.motivo else '',
            'Estado': c.estado,
            'Creación': c.fecha_creacion.strftime('%Y-%m-%d') if c.fecha_creacion else ''
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    if columnas_seleccionadas:
        disponibles = [col for col in columnas_seleccionadas if col in df.columns]
        if disponibles:
            return df[disponibles]
    return df


def dataframe_desde_contactos(contactos, columnas_seleccionadas=None):
    """
    Convierte una lista de mensajes de contacto en un DataFrame.
    Incluye información de si ya fueron respondidos.
    """
    from models.mensaje import Mensaje  # Import local para evitar circular imports
    
    data = []
    for m in contactos:
        # Verificar si existe una respuesta de admin para este hilo
        respondido = Mensaje.query.filter_by(
            tramite_folio=m.tramite_folio, 
            es_admin=True
        ).first() is not None
        
        row = {
            'Folio': m.tramite_folio,
            'Nombre': m.autor_nombre,
            'Email': m.autor_email,
            'Mensaje': m.mensaje[:150],
            'Fecha': m.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            'Respondido': 'Sí' if respondido else 'No'
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    if columnas_seleccionadas:
        disponibles = [col for col in columnas_seleccionadas if col in df.columns]
        if disponibles:
            return df[disponibles]
    return df


# ================================================================
# GENERACIÓN DE TABLAS HTML PROFESIONALES (CON ESTILOS PERSONALIZABLES)
# ================================================================

def generar_tabla_html_profesional(dataframe, titulo, estilos=None):
    """
    Genera una tabla HTML con CSS inline aplicando los estilos personalizados.
    
    Args:
        dataframe: pd.DataFrame con los datos.
        titulo: str, título de la tabla.
        estilos: dict con opciones de personalización (color_encabezado, fuente, etc.)
    
    Returns:
        str: código HTML de la tabla.
    """
    if dataframe.empty:
        return "<div class='alert alert-warning text-center'>No hay datos para mostrar.</div>"
    
    # Valores por defecto
    if estilos is None:
        estilos = {}
    
    color_encabezado = estilos.get('color_encabezado', '#2D5016')
    color_texto_encabezado = estilos.get('color_texto_encabezado', '#FFFFFF')
    color_fila_par = estilos.get('color_fila_par', '#F8F9FA')
    color_fila_impar = estilos.get('color_fila_impar', '#FFFFFF')
    fuente = estilos.get('fuente', 'Arial, sans-serif')
    tamano = estilos.get('tamano_fuente', 14)
    mostrar_bordes = estilos.get('mostrar_bordes', True)
    formato_fecha = estilos.get('formato_fecha', '%d/%m/%Y %H:%M')
    
    border_style = '1px solid #dee2e6' if mostrar_bordes else 'none'
    
    html = f"""
    <style>
        .reporte-tabla {{
            font-family: {fuente};
            font-size: {tamano}px;
            border-collapse: collapse;
            width: 100%;
        }}
        .reporte-tabla th {{
            background-color: {color_encabezado};
            color: {color_texto_encabezado};
            padding: 10px;
            border: {border_style};
        }}
        .reporte-tabla td {{
            border: {border_style};
            padding: 8px;
        }}
        .reporte-tabla tr:nth-child(even) {{
            background-color: {color_fila_par};
        }}
        .reporte-tabla tr:nth-child(odd) {{
            background-color: {color_fila_impar};
        }}
    </style>
    <div class="reporte-tabla mt-4">
        <h4 class="mb-3">{titulo}</h4>
        <div class="table-responsive">
            <table class="reporte-tabla">
                <thead>
                    <tr>
    """
    for col in dataframe.columns:
        html += f"<th>{col}</th>"
    html += """
                    </tr>
                </thead>
                <tbody>
    """
    for _, row in dataframe.iterrows():
        html += "<tr>"
        for col in dataframe.columns:
            valor = row[col]
            if isinstance(valor, (datetime, pd.Timestamp)):
                valor = valor.strftime(formato_fecha)
            html += f"<td>{valor}</td>"
        html += "</tr>"
    html += """
                </tbody>
            </table>
        </div>
    </div>
    """
    return html


# ================================================================
# EXPORTACIÓN A EXCEL CON FORMATO PROFESIONAL (CON ESTILOS)
# ================================================================

def exportar_excel_profesional(dataframe, titulo_hoja="Reporte", estilos=None):
    """
    Exporta un DataFrame a un archivo Excel en memoria (BytesIO) con formato básico.
    
    Args:
        dataframe: pd.DataFrame con los datos.
        titulo_hoja: str, nombre de la hoja (máx 31 caracteres).
        estilos: dict con opciones de personalización (color_encabezado, fuente, etc.)
    
    Returns:
        io.BytesIO: objeto tipo archivo listo para enviar.
    """
    if estilos is None:
        estilos = {}
    
    color_encabezado = estilos.get('color_encabezado', '2D5016').lstrip('#')
    color_texto = estilos.get('color_texto_encabezado', 'FFFFFF')
    fuente = estilos.get('fuente', 'Arial')
    tamano = estilos.get('tamano_fuente', 11)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name=titulo_hoja[:31], index=False)
        workbook = writer.book
        worksheet = writer.sheets[titulo_hoja[:31]]
        
        # Aplicar estilo a los encabezados
        header_font = Font(name=fuente, size=tamano+2, bold=True, color=color_texto)
        header_fill = PatternFill(start_color=color_encabezado, end_color=color_encabezado, fill_type='solid')
        for col in range(1, len(dataframe.columns)+1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_len + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    output.seek(0)
    return output


# ================================================================
# GENERACIÓN DE GRÁFICOS DE BARRAS (BASE64)
# ================================================================

def generar_grafico_barras(dataframe, columna, titulo="Gráfico de barras"):
    """
    Genera un gráfico de barras a partir de una columna categórica y lo devuelve en base64.
    
    Args:
        dataframe: pd.DataFrame.
        columna: str, nombre de la columna a agrupar.
        titulo: str, título del gráfico.
    
    Returns:
        str: imagen en formato base64 (para incrustar en HTML) o None si no es posible.
    """
    if dataframe.empty or columna not in dataframe.columns:
        return None
    
    conteo = dataframe[columna].value_counts()
    
    plt.figure(figsize=(8, 5))
    conteo.plot(kind='bar', color='#2D5016', edgecolor='black')
    plt.title(titulo, fontsize=14, fontweight='bold')
    plt.xlabel(columna, fontsize=12)
    plt.ylabel("Cantidad", fontsize=12)
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=100)
    plt.close()
    buffer.seek(0)
    
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return imagen_base64


# ================================================================
# GENERACIÓN DE PDF DESDE HTML (con WeasyPrint)
# ================================================================

def generar_pdf_desde_html(html_string):
    """
    Convierte una cadena HTML en un archivo PDF y lo devuelve en memoria.
    
    Args:
        html_string: str, código HTML completo (incluyendo estilos).
    
    Returns:
        io.BytesIO: objeto tipo archivo PDF listo para descargar.
    """
    pdf_bytes = HTML(string=html_string).write_pdf()
    output = io.BytesIO(pdf_bytes)
    output.seek(0)
    return output


def generar_pdf_desde_template(template_name, context, app=None):
    """
    Renderiza una plantilla HTML de Flask y la convierte a PDF.
    
    Args:
        template_name: str, nombre de la plantilla (ej: 'admin/reporte_pdf.html')
        context: dict, variables para la plantilla.
        app: instancia de Flask (opcional, si se llama desde un contexto de aplicación).
    
    Returns:
        io.BytesIO: objeto PDF.
    """
    if app:
        with app.app_context():
            from flask import render_template
            html = render_template(template_name, **context)
    else:
        # Asume que ya estamos dentro de un contexto de aplicación Flask
        from flask import render_template
        html = render_template(template_name, **context)
    
    return generar_pdf_desde_html(html)